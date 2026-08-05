import streamlit as st
import requests
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import time
import re
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from bs4 import BeautifulSoup
import os
import json
import hashlib
from datetime import datetime, timedelta
import uuid
import random

# ─── ANTI-DÉTECTION CONFIG ────────────────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/120.0.0.0 Safari/537.36",
]

REFERERS = [
    "https://www.google.fr/",
    "https://www.bing.com/",
    "https://duckduckgo.com/",
    "https://www.qwant.com/",
    "",
]

def get_headers_aleatoires():
    """Génère des headers aléatoires réalistes."""
    ua = random.choice(USER_AGENTS)
    referer = random.choice(REFERERS)
    headers = {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": random.choice(["fr-FR,fr;q=0.9,en;q=0.8", "fr-FR,fr;q=0.8,en-US;q=0.5", "fr,en;q=0.9"]),
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none" if not referer else "cross-site",
        "Cache-Control": "max-age=0",
    }
    if referer:
        headers["Referer"] = referer
    return headers

def delai_humain(min_s=1.5, max_s=4.5):
    """Pause aléatoire pour simuler un comportement humain."""
    time.sleep(random.uniform(min_s, max_s))

# Session persistante avec cookies
_session = None
# ─── SITES EXCLUS COMME SITE OFFICIEL ────────────────────────────────────────
SITES_EXCLUS_OFFICIEL = [
    # Annuaires FR
    "pagesjaunes.fr", "societe.com", "verif.com", "infogreffe.fr",
    "kompass.com", "europages.fr", "cylex.fr", "hoodspot.fr",
    "pappers.fr", "infobel.com", "118712.fr", "tel.fr",
    "annuaire-entreprises.data.gouv.fr", "societe.ninja",
    # Réseaux sociaux
    "facebook.com", "instagram.com", "linkedin.com", "x.com", "twitter.com",
    # Avis / comparateurs
    "tripadvisor.fr", "trustpilot.com", "yelp.fr", "lafourchette.com",
    # Cartographie
    "mappy.com", "mappy.fr", "google.com", "google.fr",
    "googleusercontent.com", "bing.com",
    # CDN / infra
    "cloudflare.com", "cloudflareinsights.com",
    # Emploi / recrutement
    "indeed.com", "indeed.fr", "pole-emploi.fr", "apec.fr",
    "hellowork.com", "welcometothejungle.com",
    # Divers
    "opencorporates.com", "dnb.com", "similarweb.com",
    "yahoo.fr", "leboncoin.fr",
]

# Domaines hébergeurs/registrars à pénaliser dans les coordonnées
DOMAINES_HEBERGEURS = [
    "gandi", "ovh", "hostmaster", "abuse", "registrar",
    "nic.", "whois", "ionos", "o2switch", "infomaniak",
    "1and1", "godaddy", "namecheap", "cloudflare", "hetzner",
    "online.net", "scaleway", "sogenactif", "support@support",
]

EMAILS_HEBERGEURS_PATTERNS = [
    r".*@.*gandi\.",
    r".*@.*ovh\.",
    r"hostmaster@",
    r"abuse@",
    r"support@support\.",
    r".*@.*registrar\.",
    r"whois@",
    r"noc@",
    r"postmaster@",
]

TELS_HEBERGEURS = [
    # Gandi
    "0970708184", "33970708184", "970708184",
    # OVH
    "0800804182", "33800804182",
    # INPI (hotline enregistrement brevets — présent dans SIRENE)
    "0890160000", "33890160000", "890160000",
    # Cloudflare / hébergeurs
    "18005550000",
    # Note : les préfixes courts (0899, 0892, 0891) sont gérés
    # par PREFIXES_SURTAXES ci-dessous (évite les faux positifs)
]

# Préfixes surtaxés à rejeter systématiquement (0890, 0891, 0892, 0899...)
PREFIXES_SURTAXES = ["0890", "0891", "0892", "0897", "0898", "0899"]

MOTS_ANNUAIRE = [
    "annuaire", "comparateur", "marketplace", "directory", "listing",
    "trouve", "trouver", "recherche", "search", "devis", "gratuit",
    "avis", "review", "rating", "score", "top10", "meilleur",
]

ACTIVITES_MOTS_APE = {
    "43": ["btp", "bâtiment", "travaux", "construction", "rénovation", "chantier", "plaquiste", "carrelage", "peinture", "plomberie", "électricité"],
    "81": ["paysage", "jardin", "espaces verts", "entretien", "tonte", "élagage", "arboriste", "pelouse"],
    "41": ["construction", "maison", "bâtiment", "logement", "immobilier", "promotion"],
    "45": ["garage", "automobile", "véhicule", "mécanique", "réparation", "carrosserie"],
    "25": ["métallerie", "serrurerie", "menuiserie", "métal", "acier", "forge"],
    "16": ["bois", "menuiserie", "charpente", "parquet", "scierie"],
    "33": ["maintenance", "réparation", "entretien", "installation"],
    "71": ["architecture", "bureau d'études", "ingénierie", "conseil"],
    "56": ["restaurant", "traiteur", "cuisine", "brasserie", "café"],
    "47": ["commerce", "vente", "magasin", "boutique", "distribution"],
}

def scorer_domaine(url, nom_entreprise, ville, code_ape, html_content=None):
    """
    Score un domaine pour déterminer s'il est le site officiel de l'entreprise.
    Retourne (score, raisons_positives, raisons_negatives, reject_site)
    reject_site=True si aucun élément de validation trouvé.
    """
    if not url:
        return 0, [], ["URL vide"], True

    score = 0
    raisons_pos = []
    raisons_neg = []
    elements_validation = 0  # Doit être >= 1 pour accepter

    # Extraire le domaine
    domain = re.sub(r"https?://", "", url).split("/")[0].lower().strip("www.")

    # ── CRITÈRES NÉGATIFS ÉLIMINATOIRES ──────────────────────────────────────
    for exclu in SITES_EXCLUS_OFFICIEL:
        if exclu in domain:
            return -100, [], [f"Site exclu : {exclu}"], True

    for mot in MOTS_ANNUAIRE:
        if mot in domain:
            raisons_neg.append(f"Domaine annuaire/comparateur : '{mot}'")
            score -= 30

    # ── CRITÈRES POSITIFS SUR LE DOMAINE ─────────────────────────────────────
    nom_clean = re.sub(r"[^a-z0-9]", "", nom_entreprise.lower())
    mots_nom = [m for m in re.sub(r"[^a-z0-9\s]", "", nom_entreprise.lower()).split() if len(m) > 2]
    domain_clean = re.sub(r"[^a-z0-9]", "", domain)
    ville_clean = re.sub(r"[^a-z0-9]", "", ville.lower())

    # Nom exact dans le domaine
    if nom_clean and nom_clean in domain_clean:
        score += 40
        raisons_pos.append("Nom exact dans le domaine")
        elements_validation += 1
    else:
        mots_trouves = [m for m in mots_nom if len(m) > 3 and m in domain_clean]
        if len(mots_trouves) >= 2:
            score += 25
            raisons_pos.append(f"Mots du nom dans le domaine : {mots_trouves}")
            elements_validation += 1
        elif len(mots_trouves) == 1:
            score += 10
            raisons_pos.append(f"Un mot du nom dans le domaine : {mots_trouves}")

    # Ville dans le domaine
    if ville_clean and len(ville_clean) > 3 and ville_clean in domain_clean:
        score += 10
        raisons_pos.append("Ville dans le domaine")
        elements_validation += 1

    # Extension .fr
    if domain.endswith(".fr"):
        score += 10
        raisons_pos.append("Extension .fr")

    # ── ANALYSE DU CONTENU HTML ───────────────────────────────────────────────
    if html_content:
        texte = html_content.get_text().lower() if hasattr(html_content, "get_text") else str(html_content).lower()

        # Présence du nom de l'entreprise
        nom_lower = nom_entreprise.lower()
        if len(nom_lower) > 3 and nom_lower in texte:
            score += 30
            raisons_pos.append("Nom exact dans le contenu")
            elements_validation += 1
        else:
            mots_trouves_texte = [m for m in mots_nom if len(m) > 3 and m in texte]
            if len(mots_trouves_texte) >= 2:
                score += 15
                raisons_pos.append(f"Mots du nom dans le contenu : {mots_trouves_texte}")
                elements_validation += 1

        # Présence SIREN/SIRET
        if re.search(r"\d{9}|\d{14}", texte):
            score += 20
            raisons_pos.append("SIREN/SIRET présent")
            elements_validation += 1

        # Présence mentions légales
        if any(x in texte for x in ["mentions légales", "mention légale", "mentions legales", "mentions-légales"]):
            score += 15
            raisons_pos.append("Mentions légales présentes")
            elements_validation += 1

        # Présence page contact
        if any(x in texte for x in ["contact", "nous contacter", "contactez", "formulaire de contact"]):
            score += 10
            raisons_pos.append("Section contact présente")

        # Présence ville dans le contenu
        if ville and len(ville) > 3 and ville.lower() in texte:
            score += 10
            raisons_pos.append("Ville présente dans le contenu")
            elements_validation += 1

        # Validation activité via code APE
        ape_prefix = code_ape[:2] if code_ape else ""
        mots_activite = ACTIVITES_MOTS_APE.get(ape_prefix, [])
        if any(mot in texte for mot in mots_activite):
            score += 10
            raisons_pos.append("Activité cohérente avec le code APE")
            elements_validation += 1

        # Critères négatifs sur le contenu
        if any(x in texte for x in ["annuaire", "liste des entreprises", "comparer les devis", "trouvez un professionnel", "nos partenaires"]):
            score -= 20
            raisons_neg.append("Contenu d'annuaire détecté")
            elements_validation = max(0, elements_validation - 1)

    # ── RÈGLE DE VALIDATION ───────────────────────────────────────────────────
    # Si on a du contenu HTML : exiger au moins 1 élément de validation
    # Si pas de contenu HTML (scraping échoué) : accepter si score domaine >= 20
    if html_content is not None:
        reject_site = elements_validation < 1
        if reject_site:
            raisons_neg.append("Rejeté : aucun élément de validation dans le contenu")
    else:
        # Sans contenu HTML, on se base uniquement sur le domaine
        score_domaine_seul = sum(
            v for v in [
                40 if (nom_clean and nom_clean in domain_clean) else 0,
                25 if len([m for m in mots_nom if len(m) > 3 and m in domain_clean]) >= 2 else 0,
                10 if (ville_clean and len(ville_clean) > 3 and ville_clean in domain_clean) else 0,
                10 if domain.endswith(".fr") else 0,
            ]
        )
        reject_site = score_domaine_seul < 20
        if reject_site:
            raisons_neg.append(f"Rejeté (sans HTML) : score domaine {score_domaine_seul} < 20")
        else:
            raisons_pos.append(f"Accepté provisoirement sans HTML (score domaine={score_domaine_seul})")

    return score, raisons_pos, raisons_neg, reject_site


def valider_site_officiel(url, nom_entreprise, ville, code_ape, soup_existant=None):
    """
    Valide et score un site web candidat.
    Retourne (url, score, confidence_label, raison)
    Un site est rejeté si reject_site=True (aucun élément de validation).
    soup_existant : si fourni, évite un double scraping.
    """
    if not url:
        return None, 0, "non_trouve", "Aucune URL"

    # Vérification rapide domaine exclu
    domain = re.sub(r"https?://", "", url).split("/")[0].lower()
    for exclu in SITES_EXCLUS_OFFICIEL:
        if exclu in domain:
            return None, -100, "exclu", f"Site exclu : {exclu}"

    # Utiliser le soup fourni ou scraper une seule fois
    soup = soup_existant if soup_existant is not None else scraper_intelligent(url, timeout=8)

    # Scorer avec validation stricte
    score, raisons_pos, raisons_neg, reject_site = scorer_domaine(
        url, nom_entreprise, ville, code_ape, soup)

    # Déterminer le niveau de confiance
    if reject_site:
        confidence = "rejeté"
    elif score >= 70:
        confidence = "élevée"
    elif score >= 40:
        confidence = "moyenne"
    elif score >= 10:
        confidence = "faible"
    else:
        confidence = "non_fiable"

    raison = ""
    if raisons_pos:
        raison += "✅ " + " | ".join(raisons_pos[:3])
    if raisons_neg:
        raison += " ❌ " + " | ".join(raisons_neg[:2])

    # Rejeter si pas d'élément de validation OU score trop bas
    if reject_site or score < 10:
        return None, score, confidence, raison

    return url, score, confidence, raison


def trouver_meilleur_site(nom_entreprise, ville, code_ape, sites_candidats):
    """
    Prend une liste de sites candidats, les score tous et retourne le meilleur.
    sites_candidats : liste de tuples (url, source) ou (url, source, soup)
    Le soup optionnel évite le double scraping (BUG 4).
    """
    meilleur_url = None
    meilleur_score = -999
    meilleur_confidence = "non_trouve"
    meilleure_raison = "Aucun site trouvé"

    for candidat in sites_candidats:
        # Supporter (url, source) et (url, source, soup)
        if len(candidat) == 3:
            url, source, soup_existant = candidat
        else:
            url, source = candidat
            soup_existant = None

        if not url:
            continue

        # Vérification rapide domaine exclu — sans scraping
        domain = re.sub(r"https?://", "", url).split("/")[0].lower()
        if any(exclu in domain for exclu in SITES_EXCLUS_OFFICIEL):
            log_enrichissement(nom_entreprise, "site_exclu", url, source)
            continue

        # Passer le soup existant pour éviter le double scraping
        url_validee, score, confidence, raison = valider_site_officiel(
            url, nom_entreprise, ville, code_ape, soup=soup_existant)

        log_enrichissement(nom_entreprise, "site_candidat_score",
                           f"{url} → {score}", f"{source}/{confidence}")

        if score > meilleur_score:
            meilleur_score = score
            meilleur_url = url_validee
            meilleur_confidence = confidence
            meilleure_raison = f"[{source}] {raison}"

        # Délai seulement si on a dû scraper (soup_existant=None)
        if soup_existant is None:
            delai_humain(0.3, 0.8)

    return meilleur_url, meilleur_confidence, meilleure_raison


# ─── SYSTÈME DE LOGS ────────────────────────────────────────────────────────

_logs_enrichissement = []  # Log en mémoire pour la session

def log_enrichissement(nom_entreprise, action, detail, source="", score=None, rejete=False):
    """Ajoute une entrée de log pour le suivi de l'enrichissement."""
    entry = {
        "entreprise": nom_entreprise[:40],
        "action": action,
        "detail": detail[:100] if detail else "",
        "source": source[:50] if source else "",
        "score": score,
        "rejete": rejete,
        "ts": time.strftime("%H:%M:%S"),
    }
    _logs_enrichissement.append(entry)
    # Limiter à 500 entrées
    if len(_logs_enrichissement) > 500:
        _logs_enrichissement.pop(0)

def get_logs():
    return _logs_enrichissement.copy()

def clear_logs():
    _logs_enrichissement.clear()


def get_session():
    global _session
    if _session is None:
        import requests as _requests
        _session = _requests.Session()
        _session.headers.update(get_headers_aleatoires())
    return _session

def requete_humaine(url, timeout=10, use_session=True):
    """Fait une requête HTTP en simulant un comportement humain."""
    try:
        headers = get_headers_aleatoires()
        if use_session:
            sess = get_session()
            sess.headers.update(headers)
            resp = sess.get(url, timeout=timeout, allow_redirects=True)
        else:
            resp = requests.get(url, timeout=timeout, headers=headers, allow_redirects=True)
        return resp
    except Exception:
        return None

# ─── PLAYWRIGHT / CAMOUFOX CONFIG ─────────────────────────────────────────────
PLAYWRIGHT_DISPONIBLE = False
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_DISPONIBLE = True
except ImportError:
    pass

CAMOUFOX_DISPONIBLE = False
try:
    from camoufox.sync_api import Camoufox
    CAMOUFOX_DISPONIBLE = True
except ImportError:
    pass

def scraper_avec_camoufox(url, timeout=15000):
    """Scrape une URL avec Camoufox - contourne les protections anti-bot avancées."""
    if not CAMOUFOX_DISPONIBLE:
        return None
    try:
        with Camoufox(headless=True) as fox:
            page = fox.new_page()
            page.set_extra_http_headers(get_headers_aleatoires())
            page.goto(url, timeout=timeout, wait_until="domcontentloaded")
            # Simuler comportement humain
            page.mouse.move(random.randint(100, 500), random.randint(100, 400))
            time.sleep(random.uniform(0.5, 1.5))
            page.mouse.wheel(0, random.randint(200, 600))
            time.sleep(random.uniform(0.3, 1.0))
            html = page.content()
            page.close()
            return BeautifulSoup(html, "lxml")
    except Exception:
        return None

def scraper_avec_playwright(url, timeout=15000):
    """Scrape une URL avec Playwright - fallback si Camoufox indisponible."""
    if not PLAYWRIGHT_DISPONIBLE:
        return None
    try:
        with sync_playwright() as p:
            browser = p.firefox.launch(headless=True)
            context = browser.new_context(
                user_agent=random.choice(USER_AGENTS),
                viewport={"width": random.randint(1200, 1920), "height": random.randint(700, 1080)},
                locale="fr-FR",
                timezone_id="Europe/Paris",
            )
            page = context.new_page()
            page.goto(url, timeout=timeout, wait_until="domcontentloaded")
            time.sleep(random.uniform(0.5, 2.0))
            html = page.content()
            browser.close()
            return BeautifulSoup(html, "lxml")
    except Exception:
        return None

def scraper_intelligent(url, timeout=10):
    """Essaie d'abord requests simple, puis Camoufox, puis Playwright si bloqué."""
    if not url:
        return None
    try:
        if not url.startswith("http"):
            url = "https://" + url
        # 1. Essai simple avec session + headers humains
        resp = requete_humaine(url, timeout=timeout)
        if resp and resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "lxml")
            # Vérifier si on a été bloqué (page captcha/cloudflare vide)
            texte = soup.get_text()
            if len(texte) > 200 and not any(x in texte.lower() for x in ["captcha", "cloudflare", "access denied", "403 forbidden", "bot detected"]):
                return soup

        # 2. Si bloqué → Camoufox
        delai_humain(1, 3)
        soup = scraper_avec_camoufox(url)
        if soup:
            return soup

        # 3. Fallback → Playwright
        delai_humain(1, 2)
        return scraper_avec_playwright(url)

    except Exception:
        return None

# ─── SUPABASE CONFIG ──────────────────────────────────────────────────────────
SUPABASE_URL = "https://ilezjfqgfjdismxqnzjo.supabase.co"
SUPABASE_KEY = "sb_publishable_FqF6D8AL0vTb_X-oJdZ9QA_1bbM3aUM"

def get_supabase():
    try:
        from supabase import create_client
        return create_client(SUPABASE_URL, SUPABASE_KEY)
    except Exception:
        return None

st.set_page_config(page_title="🎓 Outil Recherche de Stage", layout="wide")

# ─── SESSION STATE ────────────────────────────────────────────────────────────
if "entreprises" not in st.session_state:
    st.session_state.entreprises = []
if "profil_connecte" not in st.session_state:
    st.session_state.profil_connecte = None
if "relances_verifiees" not in st.session_state:
    st.session_state.relances_verifiees = False
if "relances_en_attente" not in st.session_state:
    st.session_state.relances_en_attente = []
if "historique_cache" not in st.session_state:
    st.session_state.historique_cache = {}
if "profil_cache" not in st.session_state:
    st.session_state.profil_cache = {}

def get_historique_cached(email):
    """Charge l'historique avec cache session — évite les appels Supabase répétés."""
    if email not in st.session_state.historique_cache:
        st.session_state.historique_cache[email] = charger_historique(email)
    return st.session_state.historique_cache[email]

def invalider_cache_historique(email):
    """Invalide le cache après une modification."""
    st.session_state.historique_cache.pop(email, None)

def sauvegarder_historique_cached(email, historique):
    """Sauvegarde et met à jour le cache."""
    sauvegarder_historique(email, historique)
    st.session_state.historique_cache[email] = historique
if "historique_cache" not in st.session_state:
    st.session_state.historique_cache = {}
if "profil_cache" not in st.session_state:
    st.session_state.profil_cache = {}
if "historique_cache" not in st.session_state:
    st.session_state.historique_cache = {}

# ─── CONSTANTES ───────────────────────────────────────────────────────────────
TRANCHES_SALARIES = {
    "NN": 0, "00": 0, "01": 2, "02": 5, "03": 10, "11": 15,
    "12": 25, "21": 50, "22": 75, "31": 150, "32": 250,
    "41": 375, "42": 750, "51": 1500, "52": 3500, "53": 7500
}
COLONNES_KANBAN = ["📋 À contacter", "📤 Contactée", "📬 Réponse reçue", "🤝 Entretien", "✅ Accepté", "❌ Refus"]
SUJET_RELANCE_DEFAULT = "Relance – Demande de stage d'immersion – {nom_entreprise}"
CORPS_RELANCE_DEFAULT = """Madame, Monsieur,

Je me permets de revenir vers vous suite à mon précédent email du {date_premier_contact}, dans lequel je sollicitais un stage d'immersion de deux semaines au sein de votre entreprise {nom_entreprise}.

N'ayant pas eu de retour de votre part, je souhaitais renouveler ma candidature et vous confirmer ma motivation intacte pour découvrir les métiers de votre secteur.

Je reste entièrement disponible pour en discuter à votre convenance.

Dans l'attente de votre retour, je vous adresse mes sincères salutations.

{prenom_nom}"""

# ─── FONCTIONS PROFIL ─────────────────────────────────────────────────────────
def get_profil_id(email):
    return hashlib.md5(email.strip().lower().encode()).hexdigest()

def get_historique_path(email):
    return f"historique_{get_profil_id(email)}.json"

def get_profil_path(email):
    return f"profil_{get_profil_id(email)}.json"

def charger_historique(email, force_reload=False):
    """Charge l'historique avec cache session pour éviter les appels répétés."""
    cache_key = f"hist_{get_profil_id(email)}"
    # Utiliser le cache session si disponible
    if not force_reload and cache_key in st.session_state.get("historique_cache", {}):
        return st.session_state.historique_cache[cache_key]

    profil_id = get_profil_id(email)
    historique = {}
    # Essayer Supabase d'abord
    try:
        sb = get_supabase()
        if sb:
            res = sb.table("historique").select("nom, donnees").eq("profil_id", profil_id).execute()
            if res.data:
                historique = {row["nom"]: row["donnees"] for row in res.data}
    except Exception:
        pass
    # Fallback local
    if not historique:
        path = get_historique_path(email)
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                historique = json.load(f)
    # Mettre en cache
    if "historique_cache" in st.session_state:
        st.session_state.historique_cache[cache_key] = historique
    return historique

def sauvegarder_historique(email, historique):
    profil_id = get_profil_id(email)
    # Sauvegarder sur Supabase
    try:
        sb = get_supabase()
        if sb:
            for cle, donnees in historique.items():
                sb.table("historique").upsert({
                    "id": f"{profil_id}_{cle[:50]}",
                    "profil_id": profil_id,
                    "nom": cle,
                    "donnees": donnees
                }).execute()
            # Supprimer les entrées qui ne sont plus dans l'historique
            res = sb.table("historique").select("nom").eq("profil_id", profil_id).execute()
            if res.data:
                noms_supabase = {row["nom"] for row in res.data}
                noms_locaux = set(historique.keys())
                a_supprimer = noms_supabase - noms_locaux
                for nom in a_supprimer:
                    sb.table("historique").delete().eq("profil_id", profil_id).eq("nom", nom).execute()
    except Exception:
        pass
    # Sauvegarder aussi en local comme backup
    with open(get_historique_path(email), "w", encoding="utf-8") as f:
        json.dump(historique, f, ensure_ascii=False, indent=2)

def charger_profil(email):
    profil_id = get_profil_id(email)
    # Essayer Supabase d'abord
    try:
        sb = get_supabase()
        if sb:
            res = sb.table("profils").select("donnees").eq("id", profil_id).execute()
            if res.data:
                return res.data[0]["donnees"]
    except Exception:
        pass
    # Fallback local
    path = get_profil_path(email)
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def sauvegarder_profil(email, donnees):
    profil_id = get_profil_id(email)
    # Sauvegarder sur Supabase
    try:
        sb = get_supabase()
        if sb:
            sb.table("profils").upsert({
                "id": profil_id,
                "email": email,
                "donnees": donnees
            }).execute()
    except Exception:
        pass
    # Sauvegarder aussi en local comme backup
    with open(get_profil_path(email), "w", encoding="utf-8") as f:
        json.dump(donnees, f, ensure_ascii=False, indent=2)

def charger_modeles(email):
    return charger_profil(email).get("modeles_email", {})

def sauvegarder_modeles(email, modeles):
    profil = charger_profil(email)
    profil["modeles_email"] = modeles
    sauvegarder_profil(email, profil)

def marquer_contactees(email, entreprises_envoyees):
    historique = charger_historique(email)
    for e in entreprises_envoyees:
        cle = e["nom"].strip().lower()
        if cle not in historique:
            historique[cle] = {
                "id": str(uuid.uuid4()),
                "nom": e["nom"], "ville": e["ville"],
                "email": e["email"], "telephone": e.get("telephone", ""),
                "adresse": e.get("adresse", ""), "code_ape": e.get("code_ape", ""),
                "site_web": e.get("site_web", ""), "pages_jaunes": e.get("pages_jaunes", ""),
                "facebook": e.get("facebook", ""),
                "date_contact": datetime.now().strftime("%d/%m/%Y %H:%M"),
                "date_contact_iso": datetime.now().isoformat(),
                "nb_relances": 0, "date_derniere_relance": "",
                "statut_kanban": "📤 Contactée", "repondu": False,
                "favori": False, "notes": "",
            }
        else:
            historique[cle].update({
                "email": e["email"], "site_web": e.get("site_web", ""),
                "pages_jaunes": e.get("pages_jaunes", ""), "facebook": e.get("facebook", ""),
            })
    sauvegarder_historique(email, historique)

def tranche_vers_effectif(tranche):
    return TRANCHES_SALARIES.get(str(tranche), 0)

# ─── FONCTIONS EMAIL ──────────────────────────────────────────────────────────
def envoyer_email(expediteur, mot_de_passe, destinataire, sujet, corps, chemin_cv=None):
    try:
        msg = MIMEMultipart()
        msg["From"] = expediteur
        msg["To"] = destinataire
        msg["Subject"] = sujet
        msg.attach(MIMEText(corps, "plain", "utf-8"))
        if chemin_cv and os.path.exists(chemin_cv):
            with open(chemin_cv, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(chemin_cv)}")
                msg.attach(part)
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(expediteur, mot_de_passe)
            server.sendmail(expediteur, destinataire, msg.as_string())
        return True
    except Exception as e:
        return str(e)

# ─── FONCTIONS SCRAPING ───────────────────────────────────────────────────────
# ─── PAGES DE CONTACT À EXPLORER ────────────────────────────────────────────
PAGES_CONTACT = [
    "/contact", "/contactez-nous", "/nous-contacter", "/contact.html",
    "/contact-us", "/nous-joindre",
    "/mentions-legales", "/mentions-légales", "/legal", "/legal.html",
    "/about", "/a-propos", "/qui-sommes-nous", "/notre-entreprise",
    "/support", "/service-client", "/sav", "/aide",
    "/equipe", "/team", "/notre-equipe",
    "/cgv", "/conditions-generales", "/coordonnees",
]

DOMAINES_ANNUAIRES = [
    "pagesjaunes", "facebook", "yelp", "societe", "kompass",
    "infobel", "118712", "cylex", "hoodspot", "europages",
    "mappy", "pappers", "infogreffe", "verif.com", "linkedin",
    "instagram", "twitter", "x.com", "tripadvisor", "trustpilot",
]

def extraire_contact_site(url):
    """
    Extrait email ET téléphone depuis une URL.
    Workflow :
      1. requests + extraction complète avec scoring qualité
      2. Exploration pages de contact (sitemap + pages connues)
      3. Analyse PDFs pertinents
      4. Playwright fallback si toujours pas d'email
    """
    if not url:
        return "", "", []
    try:
        if not url.startswith("http"):
            url = "https://" + url

        base = url.rstrip("/")
        est_annuaire = any(x in base for x in DOMAINES_ANNUAIRES)

        # ── ÉTAPE 1 : Scraper la page principale ─────────────────────────────
        soup = scraper_intelligent(url)
        email, telephone = "", ""
        if soup:
            email, telephone = extraire_depuis_soup(soup)

        if est_annuaire:
            return email, telephone, []

        # ── ÉTAPE 2 : Explorer les pages de contact ───────────────────────────
        if not email or not telephone:
            for path in PAGES_CONTACT:
                if email and telephone:
                    break
                try:
                    delai_humain(0.3, 0.8)
                    soup2 = scraper_intelligent(base + path, timeout=6)
                    if not soup2:
                        continue
                    em2, tel2 = extraire_depuis_soup(soup2)
                    if not email and em2:
                        email = em2
                    if not telephone and tel2:
                        telephone = tel2
                except Exception:
                    continue

        # ── ÉTAPE 3 : Analyse PDFs ────────────────────────────────────────────
        if (not email or not telephone) and soup:
            try:
                em_p, tel_p, _, _ = analyser_pdfs_site(soup, base)
                if em_p and not email:
                    email = em_p
                if tel_p and not telephone:
                    telephone = tel_p
            except Exception:
                pass

        # ── ÉTAPE 4 : Playwright fallback (si email toujours absent) ──────────
        if not email and (CAMOUFOX_DISPONIBLE or PLAYWRIGHT_DISPONIBLE):
            try:
                soup_js, emails_js = extraire_avec_playwright_js(url)
                if emails_js:
                    em_js, _ = meilleur_email(emails_js)
                    if em_js and not email:
                        email = em_js
                if soup_js and not email:
                    em2, tel2 = extraire_depuis_soup(soup_js)
                    if em2 and not email:
                        email = em2
                    if tel2 and not telephone:
                        telephone = tel2
            except Exception:
                pass

        return email, telephone, []
    except Exception:
        return "", "", []

# ─── ANALYSE PDF ─────────────────────────────────────────────────────────────

MOTS_CLES_PDF = [
    "mention", "legal", "contact", "brochure", "catalogue",
    "plaquette", "presentation", "coordonnee", "cgv",
]
_cache_pdf = {}
MAX_PDF_PAR_SITE = 3

PYMUPDF_DISPONIBLE = False
try:
    import fitz
    PYMUPDF_DISPONIBLE = True
except ImportError:
    pass

def scorer_pdf_pertinence(url_pdf):
    url_lower = url_pdf.lower()
    return sum((len(MOTS_CLES_PDF) - i) * 5
               for i, mot in enumerate(MOTS_CLES_PDF) if mot in url_lower)

def detecter_pdfs_pertinents(soup, base_url, max_pdfs=MAX_PDF_PAR_SITE):
    if not soup:
        return []
    candidats = []
    base = base_url.rstrip("/")
    domain = re.sub(r"https?://", "", base).split("/")[0]
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href.lower().endswith(".pdf"):
            continue
        if href.startswith("/"):
            href = base + href
        elif href.startswith("http"):
            if domain not in href:
                continue
        else:
            href = base + "/" + href
        candidats.append((href, scorer_pdf_pertinence(href)))
    candidats.sort(key=lambda x: x[1], reverse=True)
    return [u for u, _ in candidats[:max_pdfs]]

def extraire_texte_pdf(contenu_bytes):
    if PYMUPDF_DISPONIBLE:
        try:
            import fitz, io
            doc = fitz.open(stream=contenu_bytes, filetype="pdf")
            texte = "".join(doc[i].get_text()
                            for i in range(min(doc.page_count, 5)))
            doc.close()
            if texte.strip():
                return texte
        except Exception:
            pass
    try:
        texte_brut = contenu_bytes.decode("latin-1", errors="ignore")
        return " ".join(re.findall(r"[ -~À-ÿ]{4,}", texte_brut))
    except Exception:
        return ""

def analyser_pdf(url_pdf):
    if url_pdf in _cache_pdf:
        em, tel, _ = _cache_pdf[url_pdf]
        return em, tel, f"pdf_cache/{url_pdf.split('/')[-1][:20]}", "élevée"
    try:
        resp = requete_humaine(url_pdf, timeout=10)
        if not resp or resp.status_code != 200:
            return "", "", "", ""
        contenu = resp.content[:5 * 1024 * 1024]
        texte = extraire_texte_pdf(contenu)
        if not texte:
            return "", "", "", ""
        emails_raw = [e for e, _ in deobfusquer_emails(texte)
                      if not est_email_hebergeur(e)]
        email = emails_raw[0] if emails_raw else ""
        tels_raw = [(t, s, c) for t, s, c in extraire_tels_depuis_texte(texte, "pdf")
                    if not est_tel_hebergeur(t)]
        telephone = tels_raw[0][0] if tels_raw else ""
        _cache_pdf[url_pdf] = (email, telephone, time.time())
        if email or telephone:
            return email, telephone, f"pdf/{url_pdf.split('/')[-1][:20]}", "élevée"
    except Exception:
        pass
    return "", "", "", ""

def analyser_pdfs_site(soup, base_url):
    """Analyse les PDFs pertinents d'un site pour en extraire contacts."""
    for url_pdf in detecter_pdfs_pertinents(soup, base_url):
        try:
            delai_humain(0.5, 1.5)
            em, tel, src, conf = analyser_pdf(url_pdf)
            if em or tel:
                return em, tel, src, conf
        except Exception:
            continue
    return "", "", "", ""


# ─── DÉTECTION FORMULAIRES ───────────────────────────────────────────────────

TYPES_FORMULAIRES = {
    "contact": ["contact", "nous-contacter", "contactez", "get-in-touch"],
    "devis":   ["devis", "quote", "estimation", "demande-de-devis"],
    "support": ["support", "aide", "help", "assistance", "faq"],
    "sav":     ["sav", "service-apres-vente", "retour", "reclamation"],
}

def detecter_type_formulaire(url, texte_page):
    url_lower = url.lower()
    texte_lower = (texte_page or "").lower()
    for type_form, mots in TYPES_FORMULAIRES.items():
        if any(m in url_lower or m in texte_lower for m in mots):
            return type_form
    return "contact"

def analyser_formulaire(soup, url):
    if not soup:
        return None
    texte = soup.get_text().lower()
    for form in soup.find_all("form"):
        fh = str(form).lower()
        if any(x in fh for x in ["login", "search", "newsletter", "register"]):
            continue
        score_form = sum([
            any(x in fh for x in ["textarea", "message"]),
            "email" in fh or "mail" in fh,
            any(x in fh for x in ["submit", "envoyer", "send"]),
        ])
        if score_form >= 2:
            return {
                "contact_form_present": True,
                "form_type": detecter_type_formulaire(url, form.get_text()),
                "page_url": url,
                "confidence": "élevée" if score_form == 3 else "moyenne",
            }
    for indice in ["formulaire de contact", "contact form", "envoyez-nous"]:
        if indice in texte:
            return {
                "contact_form_present": True,
                "form_type": detecter_type_formulaire(url, texte),
                "page_url": url,
                "confidence": "faible",
            }
    return None

def chercher_formulaires_contact(base_url, soup_homepage, pages_connues=None):
    """Cherche les formulaires de contact sur le site (max 3)."""
    formulaires, analysees = [], set()
    if soup_homepage:
        r = analyser_formulaire(soup_homepage, base_url)
        if r:
            formulaires.append(r)
        analysees.add(base_url)
    pages = [(u, s) for u, s in (pages_connues or [])
             if any(m in u.lower()
                    for mots in TYPES_FORMULAIRES.values() for m in mots)][:5]
    for page_url, _ in pages:
        if page_url in analysees or len(formulaires) >= 3:
            break
        try:
            delai_humain(0.3, 0.8)
            s = scraper_intelligent(page_url, timeout=6)
            if s:
                r = analyser_formulaire(s, page_url)
                if r:
                    formulaires.append(r)
            analysees.add(page_url)
        except Exception:
            continue
    return formulaires


# ─── PLAYWRIGHT FALLBACK ──────────────────────────────────────────────────────

def extraire_avec_playwright_js(url):
    """
    Scrape une URL avec Camoufox ou Playwright pour les sites JS.
    Retourne (soup, emails_dynamiques).
    Utilisé uniquement en fallback si email absent après requests.
    """
    if CAMOUFOX_DISPONIBLE:
        try:
            with Camoufox(headless=True) as fox:
                page = fox.new_page()
                page.set_extra_http_headers(get_headers_aleatoires())
                page.goto(url, timeout=15000, wait_until="networkidle")
                time.sleep(random.uniform(1.5, 3.0))
                page.evaluate("window.scrollTo(0, document.body.scrollHeight / 2)")
                time.sleep(random.uniform(0.5, 1.0))
                mailtos = page.evaluate("""
                    () => Array.from(document.querySelectorAll('a[href^="mailto:"]'))
                             .map(a => a.href.replace('mailto:', '').split('?')[0])
                """) or []
                html = page.content()
                page.close()
                soup = BeautifulSoup(html, "lxml")
                emails_js = [(normaliser_email(m), "mailto_js_dynamique", "élevée")
                             for m in mailtos if normaliser_email(m)
                             and not est_email_hebergeur(normaliser_email(m))]
                return soup, emails_js
        except Exception:
            pass
    if PLAYWRIGHT_DISPONIBLE:
        try:
            with sync_playwright() as p:
                browser = p.firefox.launch(headless=True)
                ctx = browser.new_context(
                    user_agent=random.choice(USER_AGENTS),
                    viewport={"width": random.randint(1200, 1920),
                              "height": random.randint(700, 1080)},
                    locale="fr-FR", timezone_id="Europe/Paris",
                )
                page = ctx.new_page()
                page.goto(url, timeout=15000, wait_until="networkidle")
                time.sleep(random.uniform(1.0, 2.0))
                mailtos = page.evaluate("""
                    () => Array.from(document.querySelectorAll('a[href^="mailto:"]'))
                             .map(a => a.href.replace('mailto:', '').split('?')[0])
                """) or []
                html = page.content()
                browser.close()
                soup = BeautifulSoup(html, "lxml")
                emails_js = [(normaliser_email(m), "mailto_js_dynamique", "élevée")
                             for m in mailtos if normaliser_email(m)
                             and not est_email_hebergeur(normaliser_email(m))]
                return soup, emails_js
        except Exception:
            pass
    return None, []


# ─── NORMALISATION ───────────────────────────────────────────────────────────
# ─── SCORING QUALITÉ COORDONNÉES ────────────────────────────────────────────

SOURCE_SCORES_EMAIL = {
    # Sources haute confiance
    "schema.org": 100,
    "JSON-LD": 100,
    "mailto:": 95,
    "données_structurées": 95,
    "mailto_js_dynamique": 90,
    "footer/direct": 85,
    "footer/mailto:": 85,
    "section_contact": 80,
    "mentions_legales": 80,
    "texte_page/direct": 60,
    # Sources basse confiance
    "pdf/": 50,
    "whois": 20,
    "reseau_social/": 40,
    "meta_tags": 30,
}

SOURCE_SCORES_TEL = {
    "schema.org": 100,
    "JSON-LD": 100,
    "balise_tel:": 95,
    "données_structurées": 95,
    "footer": 85,
    "section_contact": 80,
    "mentions_legales": 80,
    "texte_page": 60,
    "pdf/": 50,
    "reseau_social/": 40,
}

def est_email_hebergeur(email):
    """Détecte si un email appartient à un hébergeur/registrar."""
    if not email:
        return False
    email_lower = email.lower()
    import re as _re
    for pattern in EMAILS_HEBERGEURS_PATTERNS:
        if _re.match(pattern, email_lower):
            return True
    for domaine in DOMAINES_HEBERGEURS:
        if domaine in email_lower:
            return True
    return False

def est_tel_hebergeur(telephone):
    """Détecte si un téléphone appartient à un hébergeur ou est surtaxé."""
    if not telephone:
        return False
    tel_clean = re.sub(r"[^\d]", "", telephone)
    # Vérifier les numéros connus d'hébergeurs
    for t in TELS_HEBERGEURS:
        if len(t) >= 6 and t in tel_clean:  # longueur min 6 pour éviter faux positifs
            return True
    # Vérifier les préfixes surtaxés (0890, 0891...)
    for prefix in PREFIXES_SURTAXES:
        if tel_clean.startswith(prefix) or tel_clean.startswith("33" + prefix[1:]):
            return True
    return False

def score_qualite_email(email, source):
    """Retourne un score de qualité pour un email selon sa source."""
    if not email:
        return 0
    if est_email_hebergeur(email):
        return -50  # Pénalité forte
    score = 0
    for src_key, src_score in SOURCE_SCORES_EMAIL.items():
        if src_key in source:
            score = max(score, src_score)
    return score if score > 0 else 30  # Score par défaut

def score_qualite_tel(telephone, source):
    """Retourne un score de qualité pour un téléphone selon sa source."""
    if not telephone:
        return 0
    if est_tel_hebergeur(telephone):
        return -50
    score = 0
    for src_key, src_score in SOURCE_SCORES_TEL.items():
        if src_key in source:
            score = max(score, src_score)
    return score if score > 0 else 30

def meilleur_email(emails_avec_source):
    """
    Choisit le meilleur email parmi une liste (email, source, confidence).
    Élimine les emails hébergeurs, trie par score qualité.
    """
    if not emails_avec_source:
        return "", ""
    candidats = []
    for em, src, conf in emails_avec_source:
        if not em or est_email_hebergeur(em):
            continue
        score = score_qualite_email(em, src)
        candidats.append((score, em, src))
    if not candidats:
        return "", ""
    candidats.sort(reverse=True)
    return candidats[0][1], candidats[0][2]

def meilleur_tel(tels_avec_source):
    """
    Choisit le meilleur téléphone parmi une liste (tel, source, confidence).
    Élimine les téléphones hébergeurs, trie par score qualité.
    """
    if not tels_avec_source:
        return "", ""
    candidats = []
    for tel, src, conf in tels_avec_source:
        if not tel or est_tel_hebergeur(tel):
            continue
        score = score_qualite_tel(tel, src)
        candidats.append((score, tel, src))
    if not candidats:
        return "", ""
    candidats.sort(reverse=True)
    return candidats[0][1], candidats[0][2]


def normaliser_email(email):
    if not email: return ""
    import html as _html
    email = _html.unescape(str(email)).strip().lower()
    email = re.sub(r"\s+", "", email).strip(".,;:()[]{}<>")
    return email

def normaliser_tel(tel):
    if not tel: return None
    t = re.sub(r"[\s.\-()]", "", str(tel))
    t = re.sub(r"[^\d+]", "", t)
    if t.startswith("0033"): t = "+33" + t[4:]
    elif t.startswith("0") and len(t) == 10: t = "+33" + t[1:]
    elif t.startswith("+33") and len(t) == 12: pass
    elif t.startswith("+") and len(t) >= 10: pass
    else: return None
    return t

def valider_tel(t):
    if not t: return False
    chiffres = re.sub(r"[^\d]", "", t)
    if t.startswith("+33"): return len(chiffres) == 11
    return 10 <= len(chiffres) <= 15

EXCLUS_EMAIL = [".png", ".jpg", ".gif", ".svg", "sentry", "example",
                "wix", "wordpress", "jquery", "schema.org", "noreply",
                "no-reply", "donotreply", "test@", "info@info"]

TEL_PATTERN = r"(?:\(\+33\)|\+33|0033)[\s.\-]?[1-9](?:[\s.\-]?\d{2}){4}|0[1-9](?:[\s.\-]?\d{2}){4}"

def nettoyer_tel(t):
    return re.sub(r"[\s.\-]", "", t)

def deobfusquer_emails(texte):
    emails = []
    seen = set()
    patterns = [
        (r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", "direct"),
        (r"([a-zA-Z0-9._%+\-]+)\s*\[at\]\s*([a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})", "[at]"),
        (r"([a-zA-Z0-9._%+\-]+)\s*\(at\)\s*([a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})", "(at)"),
        (r"([a-zA-Z0-9._%+\-]+)\s+@\s+([a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})", "espaces"),
        (r"([a-zA-Z0-9._%+\-]+)\s+arobase\s+([a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})", "arobase"),
    ]
    for pattern, src in patterns:
        for m in re.finditer(pattern, texte, re.IGNORECASE):
            if src == "direct":
                em = normaliser_email(m.group(0))
            else:
                em = normaliser_email(f"{m.group(1)}@{m.group(2)}")
            if em and em not in seen and "@" in em and not any(x in em for x in EXCLUS_EMAIL):
                seen.add(em)
                emails.append((em, src))
    return emails

def extraire_emails_complet(soup):
    """Extrait tous les emails depuis un soup avec source et confidence."""
    if not soup: return []
    resultats = []

    # 1. Schema.org
    for tag in soup.find_all(True, itemprop="email"):
        val = tag.get("content") or tag.get_text(strip=True)
        em = normaliser_email(val)
        if em: resultats.append((em, "schema.org", "élevée"))

    # 2. JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list): data = data[0]
            em = normaliser_email(data.get("email", "") or "")
            if em: resultats.append((em, "JSON-LD", "élevée"))
        except Exception: pass

    # 3. mailto:
    for a in soup.find_all("a", href=True):
        if "mailto:" in a["href"]:
            em = normaliser_email(a["href"].replace("mailto:", "").split("?")[0])
            if em: resultats.append((em, "mailto:", "élevée"))

    # 4. Footer
    for footer in soup.find_all(["footer", "div"], class_=re.compile(r"footer|pied|bottom", re.I)):
        for em, src in deobfusquer_emails(footer.get_text()):
            resultats.append((em, f"footer/{src}", "élevée"))

    # 5. Texte complet
    for em, src in deobfusquer_emails(soup.get_text()):
        resultats.append((em, f"texte/{src}", "moyenne"))

    # Dédupliquer
    seen = {}
    ordre = ["élevée", "moyenne", "faible"]
    for em, src, conf in resultats:
        if em not in seen or ordre.index(conf) < ordre.index(seen[em][1]):
            seen[em] = (src, conf)
    return [(em, src, conf) for em, (src, conf) in seen.items()]

def extraire_tels_depuis_texte(texte, source="texte"):
    """Extrait téléphones depuis un texte."""
    resultats = []
    seen = set()
    for m in re.finditer(TEL_PATTERN, texte):
        tel = normaliser_tel(m.group(0))
        if tel and valider_tel(tel) and tel not in seen:
            chiffres = re.sub(r"[^\d]", "", tel)
            if len(set(chiffres)) >= 3:
                seen.add(tel)
                resultats.append((tel, source, "élevée"))
    return resultats

def extraire_tels_complet(soup):
    """Extrait tous les téléphones depuis un soup."""
    if not soup: return []
    resultats = []

    # Schema.org
    for tag in soup.find_all(True, itemprop="telephone"):
        val = tag.get("content") or tag.get_text(strip=True)
        tel = normaliser_tel(val)
        if tel and valider_tel(tel):
            resultats.append((tel, "schema.org", "élevée"))

    # JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list): data = data[0]
            tel = normaliser_tel(data.get("telephone", "") or "")
            if tel and valider_tel(tel):
                resultats.append((tel, "JSON-LD", "élevée"))
        except Exception: pass

    # tel:
    for a in soup.find_all("a", href=True):
        if "tel:" in a["href"]:
            tel = normaliser_tel(a["href"].replace("tel:", ""))
            if tel and valider_tel(tel):
                resultats.append((tel, "tel:", "élevée"))

    # Texte
    resultats.extend(extraire_tels_depuis_texte(soup.get_text(), "texte"))

    # Dédupliquer
    seen = {}
    ordre = ["élevée", "moyenne", "faible"]
    for tel, src, conf in resultats:
        if tel not in seen or ordre.index(conf) < ordre.index(seen[tel][1]):
            seen[tel] = (src, conf)
    return [(tel, src, conf) for tel, (src, conf) in seen.items()]

TYPES_ORGANISATION = [
    "Organization", "LocalBusiness", "Store", "ConstructionContractor",
    "HomeAndConstructionBusiness", "HousePainter", "HVACBusiness",
    "Plumber", "RoofingContractor", "GeneralContractor",
    "LandscapingService", "ProfessionalService",
]

def extraire_donnees_structurees(soup):
    """
    Extrait données depuis JSON-LD + Microdata.
    Retourne (email, telephone, nom, adresse, source, confidence)
    """
    if not soup: return "", "", "", "", "", ""

    resultats = []

    # JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            raw = re.sub(r"//[^\n]*\n", "\n", script.string or "")
            data = json.loads(raw)
            items = data if isinstance(data, list) else [data]
            for item in items:
                if "@graph" in item:
                    items.extend(item["@graph"])
                    continue
                type_item = item.get("@type", "")
                if isinstance(type_item, list): type_item = type_item[0]
                em = normaliser_email(item.get("email", "") or "")
                tel = normaliser_tel(item.get("telephone", "") or "")
                nom = item.get("name", "") or ""
                addr = item.get("address", {})
                adresse = addr if isinstance(addr, str) else " ".join(
                    str(addr.get(k, "")) for k in
                    ["streetAddress","postalCode","addressLocality"] if addr.get(k))
                if em or (tel and valider_tel(tel)):
                    conf = "élevée" if type_item in TYPES_ORGANISATION else "moyenne"
                    resultats.append((em, tel if valider_tel(tel or "") else "",
                                     nom, adresse.strip(), "JSON-LD", conf))
        except Exception: pass

    # Microdata
    for item_el in soup.find_all(True, itemscope=True):
        item_type = item_el.get("itemtype", "")
        em, tel, nom, adresse = "", "", "", ""
        for prop_el in item_el.find_all(True, itemprop=True):
            prop = prop_el.get("itemprop", "").lower()
            val = (prop_el.get("content") or
                   prop_el.get("href","").replace("mailto:","").replace("tel:","") or
                   prop_el.get_text(strip=True))
            if prop == "email": em = normaliser_email(val)
            elif prop == "telephone": tel = normaliser_tel(val) or ""
            elif prop == "name": nom = val
            elif prop in ("streetaddress","postalcode","addresslocality"):
                adresse += " " + val
        if em or (tel and valider_tel(tel)):
            conf = "élevée" if any(t in item_type for t in TYPES_ORGANISATION) else "moyenne"
            resultats.append((em, tel, nom, adresse.strip(), "Microdata", conf))

    if not resultats: return "", "", "", "", "", ""

    # Trier par confiance et complétude
    ordre = {"élevée": 0, "moyenne": 1, "faible": 2}
    resultats.sort(key=lambda r: (ordre.get(r[5], 2), -sum(1 for x in r[:4] if x)))

    # Fusionner les meilleurs
    em_f, tel_f, nom_f, addr_f, src_f, conf_f = "", "", "", "", "", ""
    for em, tel, nom, addr, src, conf in resultats:
        if not em_f and em: em_f, src_f, conf_f = em, src, conf
        if not tel_f and tel: tel_f = tel
        if not nom_f and nom: nom_f = nom
        if not addr_f and addr: addr_f = addr

    return em_f, tel_f, nom_f, addr_f, src_f, conf_f


def extraire_depuis_soup(soup):
    """
    Extrait email ET téléphone depuis un BeautifulSoup.
    Priorité : données structurées > mailto > footer > texte.
    Filtre automatiquement les coordonnées d'hébergeurs/registrars.
    """
    if not soup:
        return "", ""

    email, telephone = "", ""

    # ── 1. Données structurées JSON-LD + Microdata (priorité max) ────────────
    try:
        em_s, tel_s, _, _, _, _ = extraire_donnees_structurees(soup)
        if em_s and not est_email_hebergeur(em_s):
            email = em_s
        if tel_s and not est_tel_hebergeur(tel_s):
            telephone = tel_s
    except Exception:
        pass

    # ── 2. Emails via extraction complète + scoring qualité ──────────────────
    if not email:
        try:
            emails = extraire_emails_complet(soup)
            if emails:
                email, _ = meilleur_email(emails)
        except Exception:
            pass

    # ── 3. Téléphone via extraction complète + scoring qualité ───────────────
    if not telephone:
        try:
            tels = extraire_tels_complet(soup)
            if tels:
                telephone, _ = meilleur_tel(tels)
        except Exception:
            pass

    return email.strip() if email else "", telephone.strip() if telephone else ""


def chercher_mappy(nom, ville):
    """Cherche sur Mappy - très bien pour les artisans."""
    try:
        query = f"{nom} {ville}"
        url = f"https://fr.mappy.com/activite/{requests.utils.quote(nom.lower())}/{requests.utils.quote(ville.lower())}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        # Chercher directement le tel et email sur la page
        em, tel = extraire_depuis_soup(soup)
        if em or tel:
            return url, em, tel
        # Sinon chercher un lien vers la fiche
        lien = soup.find("a", href=re.compile(r"/poi/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://fr.mappy.com" + href
            return href, "", ""
    except Exception:
        pass
    return None, "", ""

def chercher_pappers(nom, ville):
    """Cherche sur Pappers - annuaire officiel des entreprises françaises."""
    try:
        query = f"{nom} {ville}"
        url = f"https://www.pappers.fr/recherche?q={requests.utils.quote(query)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        lien = soup.find("a", href=re.compile(r"/entreprise/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://www.pappers.fr" + href
            # Scraper la fiche entreprise
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            em, tel = extraire_depuis_soup(soup2)
            return em, tel
    except Exception:
        pass
    return "", ""

def chercher_annuaire_tel(nom, ville):
    """Cherche sur tel.fr - annuaire téléphonique professionnel."""
    try:
        query = f"{nom} {ville}"
        url = f"https://www.tel.fr/recherche/{requests.utils.quote(query)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        # Chercher le premier résultat professionnel
        for a in soup.find_all("a", href=re.compile(r"/pro/")):
            href = a["href"]
            if not href.startswith("http"):
                href = "https://www.tel.fr" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            em, tel = extraire_depuis_soup(soup2)
            if tel:
                return em, tel
    except Exception:
        pass
    return "", ""

def chercher_google_contact(nom, ville):
    """Cherche email et tel directement via Google en ciblant les résultats de contact."""
    try:
        # Cherche spécifiquement email ou contact
        for query_suffix in ["email contact", "téléphone contact", "site officiel"]:
            query = f"{nom} {ville} {query_suffix}"
            url = f"https://www.google.com/search?q={requests.utils.quote(query)}"
            resp = requete_humaine(url)
            soup = BeautifulSoup(resp.text, "lxml")
            texte = soup.get_text()
            # Chercher email dans les résultats Google directement
            emails = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", texte)
            emails = [e for e in emails if not any(x in e.lower() for x in EXCLUS_EMAIL + ["google", "goog"])]
            if emails:
                return emails[0], ""
            time.sleep(0.5)
    except Exception:
        pass
    return "", ""

def chercher_api_sirene(siren):
    """Cherche via l'API Sirene INSEE - source officielle."""
    try:
        url = f"https://api.insee.fr/entreprises/sirene/V3.11/siret?q=siren:{siren}&nombre=1"
        resp = requete_humaine(url, timeout=8)
        data = resp.json()
        etablissements = data.get("etablissements", [])
        if etablissements:
            adresse = etablissements[0].get("adresseEtablissement", {})
            tel = adresse.get("numeroVoieEtablissement", "")
            return "", tel
    except Exception:
        pass
    return "", ""

def chercher_annuaire_gouv(nom, ville):
    """Cherche via l'API Annuaire des Entreprises data.gouv.fr."""
    try:
        url = f"https://recherche-entreprises.api.gouv.fr/search?q={requests.utils.quote(nom)}&departement=&limite=3"
        resp = requete_humaine(url)
        data = resp.json()
        for r in data.get("results", []):
            if ville.lower() in r.get("siege", {}).get("libelle_commune", "").lower():
                siege = r.get("siege", {})
                tel = siege.get("telephone", "") or ""
                email = siege.get("email", "") or ""
                # Essayer aussi le site web retourné
                site = r.get("site_web", "") or ""
                if site and not email:
                    em2, tel2 = extraire_contact_site(site)
                    if em2:
                        email = em2
                    if tel2 and not tel:
                        tel = tel2
                if tel or email:
                    return email, tel
    except Exception:
        pass
    return "", ""

def chercher_infobel(nom, ville):
    """Cherche sur Infobel.fr - annuaire pro français."""
    try:
        query = f"{nom} {ville}"
        url = f"https://www.infobel.com/fr/france/search/default.aspx?kw={requests.utils.quote(query)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        em, tel = extraire_depuis_soup(soup)
        if tel or em:
            return em, tel
        # Chercher un lien vers la fiche
        lien = soup.find("a", href=re.compile(r"/fr/france/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://www.infobel.com" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_118712(nom, ville):
    """Cherche sur 118712.fr - annuaire téléphonique pro."""
    try:
        query = f"{nom} {ville}"
        url = f"https://www.118712.fr/recherche?quoi={requests.utils.quote(nom)}&ou={requests.utils.quote(ville)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        # Chercher le numéro directement
        em, tel = extraire_depuis_soup(soup)
        if tel:
            return em, tel
        # Chercher lien fiche
        lien = soup.find("a", href=re.compile(r"/fiche/|/pro/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://www.118712.fr" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_cylex(nom, ville):
    """Cherche sur Cylex.fr - annuaire entreprises locales."""
    try:
        url = f"https://www.cylex.fr/recherche/{requests.utils.quote(nom.lower())}--{requests.utils.quote(ville.lower())}.html"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        em, tel = extraire_depuis_soup(soup)
        if tel or em:
            return em, tel
        lien = soup.find("a", href=re.compile(r"/company/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://www.cylex.fr" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_hoodspot(nom, ville):
    """Cherche sur Hoodspot.fr - spécialisé artisans de proximité."""
    try:
        query = f"{nom} {ville}"
        url = f"https://hoodspot.fr/search?q={requests.utils.quote(query)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        lien = soup.find("a", href=re.compile(r"/place/|/pro/|/entreprise/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://hoodspot.fr" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_europages(nom, ville):
    """Cherche sur Europages.fr - annuaire B2B européen."""
    try:
        query = f"{nom} {ville}"
        url = f"https://www.europages.fr/entreprises/{requests.utils.quote(query)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        lien = soup.find("a", href=re.compile(r"/entreprise/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://www.europages.fr" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def deviner_emails_domaine(site_web, nom):
    """Génère et vérifie des emails probables à partir du domaine."""
    if not site_web:
        return ""
    try:
        domaine = re.sub(r'https?://', '', site_web).split('/')[0].strip()
        # Nettoyer le nom pour générer des variantes
        nom_clean = re.sub(r'[^a-zA-Z]', '', nom.lower())[:15]
        candidats = [
            f"contact@{domaine}",
            f"info@{domaine}",
            f"bonjour@{domaine}",
            f"accueil@{domaine}",
            f"devis@{domaine}",
            f"{nom_clean}@{domaine}",
        ]
        for email_candidat in candidats:
            if verifier_email_smtp(email_candidat):
                return email_candidat
    except Exception:
        pass
    return ""

def verifier_email_smtp(email):
    """Vérifie si un email existe via SMTP sans envoyer de message."""
    try:
        domaine = email.split("@")[1]
        import dns.resolver
        records = dns.resolver.resolve(domaine, 'MX')
        mx = sorted(records, key=lambda r: r.preference)[0].exchange.to_text()
        with smtplib.SMTP(timeout=5) as smtp:
            smtp.connect(mx, 25)
            smtp.helo("check.local")
            smtp.mail("check@check.local")
            code, _ = smtp.rcpt(email)
            return code == 250
    except Exception:
        pass
    return False

def chercher_google_maps(nom, ville):
    """Cherche sur Google Maps via la page publique."""
    try:
        query = f"{nom} {ville}"
        url = f"https://www.google.com/maps/search/{requests.utils.quote(query)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        em, tel = extraire_depuis_soup(soup)
        # Chercher aussi le site web dans la page
        site = ""
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("http") and "google" not in href and "maps" not in href:
                site = href
                break
        return em, tel, site
    except Exception:
        pass
    return "", "", ""

def chercher_cma(nom, ville):
    """Cherche sur l'annuaire de la Chambre des Métiers (CMA)."""
    try:
        url = f"https://www.artisanat.fr/trouver-un-artisan?search={requests.utils.quote(nom)}&location={requests.utils.quote(ville)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        em, tel = extraire_depuis_soup(soup)
        if tel or em:
            return em, tel
        lien = soup.find("a", href=re.compile(r"/artisan/|/fiche/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://www.artisanat.fr" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_qualibat(nom, ville):
    """Cherche sur Qualibat - label RGE entreprises BTP."""
    try:
        url = f"https://qualibat.com/annuaire/?search={requests.utils.quote(nom)}&city={requests.utils.quote(ville)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        lien = soup.find("a", href=re.compile(r"/entreprise/|/fiche/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://qualibat.com" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_faire_gouv(nom, ville):
    """Cherche sur faire.gouv.fr - annuaire RGE officiel."""
    try:
        url = f"https://www.faire.gouv.fr/trouver-un-professionnel?businessName={requests.utils.quote(nom)}&city={requests.utils.quote(ville)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        em, tel = extraire_depuis_soup(soup)
        if tel or em:
            return em, tel
        lien = soup.find("a", href=re.compile(r"/professionnel/|/entreprise/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://www.faire.gouv.fr" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_devis_fr(nom, ville):
    """Cherche sur Devis.fr - mise en relation artisans."""
    try:
        query = f"{nom} {ville}"
        url = f"https://www.devis.fr/annuaire/?q={requests.utils.quote(query)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        lien = soup.find("a", href=re.compile(r"/artisan/|/entreprise/|/pro/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://www.devis.fr" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_infogreffe(siren):
    """Cherche sur Infogreffe - Registre du Commerce."""
    try:
        url = f"https://www.infogreffe.fr/entreprise/{siren}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        return extraire_depuis_soup(soup)
    except Exception:
        pass
    return "", ""

def chercher_whois(site_web):
    """Cherche l'email dans les données Whois du domaine."""
    try:
        if not site_web:
            return ""
        # Extraire le domaine
        domain = re.sub(r"https?://", "", site_web).split("/")[0].strip()
        url = f"https://www.whois.com/whois/{domain}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        texte = soup.get_text()
        emails = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", texte)
        emails = [e for e in emails if not any(x in e.lower() for x in EXCLUS_EMAIL + ["whois", "abuse", "privacy"])]
        return emails[0] if emails else ""
    except Exception:
        pass
    return ""

def chercher_linkedin_dirigeant(nom_entreprise, nom_dirigeant):
    """Cherche le contact du dirigeant sur LinkedIn."""
    try:
        if not nom_dirigeant:
            return "", ""
        query = f"{nom_dirigeant} {nom_entreprise} linkedin"
        url = f"https://www.google.com/search?q={requests.utils.quote(query)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        for a in soup.find_all("a", href=True):
            if "linkedin.com/in/" in a["href"]:
                return a["href"], ""
    except Exception:
        pass
    return "", ""

def chercher_opencorporates(nom, ville):
    """Cherche sur OpenCorporates - base mondiale open source."""
    try:
        url = f"https://opencorporates.com/companies/fr?q={requests.utils.quote(nom)}&utf8=✓"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        lien = soup.find("a", href=re.compile(r"/companies/fr/"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://opencorporates.com" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_dnb(nom, ville):
    """Cherche sur Dun & Bradstreet."""
    try:
        url = f"https://www.dnb.com/business-directory/company-search.html?searchTerm={requests.utils.quote(nom)}&location={requests.utils.quote(ville)}"
        resp = requete_humaine(url)
        soup = BeautifulSoup(resp.text, "lxml")
        lien = soup.find("a", href=re.compile(r"/business-directory/company-profiles"))
        if lien:
            href = lien["href"]
            if not href.startswith("http"):
                href = "https://www.dnb.com" + href
            resp2 = requete_humaine(href)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            return extraire_depuis_soup(soup2)
    except Exception:
        pass
    return "", ""

def chercher_site_direct(nom, ville):
    """Tente de deviner directement le domaine de l'entreprise."""
    try:
        # Nettoyer le nom pour créer des variantes de domaine
        nom_clean = nom.lower()
        nom_clean = re.sub(r"[^a-z0-9\s]", "", nom_clean).strip()
        mots = nom_clean.split()
        ville_clean = ville.lower().replace(" ", "-")

        # Générer des variantes de domaine
        variantes = []
        if mots:
            base = "-".join(mots[:3])  # ex: dupont-btp-nord
            variantes += [
                f"{base}.fr", f"{base}.com",
                f"{mots[0]}.fr", f"{mots[0]}.com",
                f"{base}-{ville_clean}.fr",
                f"{mots[0]}-{ville_clean}.fr",
            ]
            if len(mots) >= 2:
                variantes += [
                    f"{mots[0]}-{mots[1]}.fr",
                    f"{mots[0]}-{mots[1]}.com",
                ]

        for domaine in variantes:
            try:
                url = f"https://{domaine}"
                resp = requete_humaine(url, timeout=5)
                if resp and resp.status_code == 200:
                    return url
            except Exception:
                continue
    except Exception:
        pass
    return None

def chercher_viewdns(nom_dirigeant):
    """Cherche les domaines via Whois inversé sur viewdns.info."""
    try:
        if not nom_dirigeant:
            return None
        url = f"https://viewdns.info/reversewhois/?q={requests.utils.quote(nom_dirigeant)}"
        soup = scraper_intelligent(url)
        if soup:
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if ".fr" in href or ".com" in href:
                    domaine = re.sub(r"https?://", "", href).split("/")[0]
                    if domaine:
                        return f"https://{domaine}"
    except Exception:
        pass
    return None

def chercher_pappers_site(nom, ville):
    """Cherche le site web via Pappers."""
    try:
        url = f"https://www.pappers.fr/recherche?q={requests.utils.quote(nom + ' ' + ville)}"
        soup = scraper_intelligent(url)
        if soup:
            lien = soup.find("a", href=re.compile(r"/entreprise/"))
            if lien:
                href = lien["href"]
                if not href.startswith("http"):
                    href = "https://www.pappers.fr" + href
                delai_humain(0.5, 1.5)
                soup2 = scraper_intelligent(href)
                if soup2:
                    # Chercher le site web dans la fiche
                    for a in soup2.find_all("a", href=True):
                        href2 = a["href"]
                        if href2.startswith("http") and "pappers" not in href2:
                            return href2
    except Exception:
        pass
    return None

def chercher_societe_site(nom, ville):
    """Cherche le site web via Societe.com."""
    try:
        url = f"https://www.societe.com/cgi-bin/search?champs={requests.utils.quote(nom + ' ' + ville)}"
        soup = scraper_intelligent(url)
        if soup:
            lien = soup.find("a", href=re.compile(r"/societe/"))
            if lien:
                href = lien["href"]
                if not href.startswith("http"):
                    href = "https://www.societe.com" + href
                delai_humain(0.5, 1.5)
                soup2 = scraper_intelligent(href)
                if soup2:
                    for a in soup2.find_all("a", href=True):
                        href2 = a["href"]
                        if href2.startswith("http") and "societe.com" not in href2:
                            return href2
    except Exception:
        pass
    return None

def chercher_kompass_site(nom, ville):
    """Cherche le site web via Kompass."""
    try:
        url = f"https://fr.kompass.com/searchCompanies?text={requests.utils.quote(nom + ' ' + ville)}"
        soup = scraper_intelligent(url)
        if soup:
            lien = soup.find("a", href=re.compile(r"/c/"))
            if lien:
                href = lien["href"]
                if not href.startswith("http"):
                    href = "https://fr.kompass.com" + href
                delai_humain(0.5, 1.5)
                soup2 = scraper_intelligent(href)
                if soup2:
                    for a in soup2.find_all("a", href=True):
                        href2 = a["href"]
                        if href2.startswith("http") and "kompass" not in href2:
                            return href2
    except Exception:
        pass
    return None

def chercher_linkedin_site(nom, ville):
    """Cherche le site web via la page LinkedIn de l'entreprise."""
    try:
        query = f"{nom} {ville} site:linkedin.com/company"
        url = f"https://www.google.com/search?q={requests.utils.quote(query)}"
        soup = scraper_intelligent(url)
        if soup:
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if "linkedin.com/company" in href:
                    if not href.startswith("http"):
                        href = href.split("/url?q=")[-1].split("&")[0]
                    delai_humain(1, 2)
                    soup2 = scraper_intelligent(href)
                    if soup2:
                        for a2 in soup2.find_all("a", href=True):
                            href2 = a2["href"]
                            if href2.startswith("http") and "linkedin" not in href2:
                                return href2
    except Exception:
        pass
    return None

def chercher_site_tentative_directe(nom, ville):
    """Tente directement des URLs probables basées sur le nom de l'entreprise."""
    try:
        nom_clean = nom.lower()
        nom_clean = re.sub(r"[^a-z0-9\s-]", "", nom_clean)
        nom_clean = re.sub(r"\s+", "-", nom_clean.strip())
        mots = nom_clean.split("-")[:3]
        nom_court = "-".join(mots)
        ville_clean = re.sub(r"[^a-z0-9]", "", ville.lower())

        candidats = [
            f"https://www.{nom_court}.fr",
            f"https://www.{nom_court}.com",
            f"https://{nom_court}.fr",
            f"https://{nom_court}.com",
            f"https://www.{nom_court}-{ville_clean}.fr",
            f"https://{nom_court}-{ville_clean}.fr",
        ]
        for url in candidats:
            try:
                # Utiliser requests direct avec timeout court pour la tentative
                resp = requests.get(url, timeout=5, headers=get_headers_aleatoires(),
                                    allow_redirects=True)
                if resp and resp.status_code == 200 and len(resp.text) > 500:
                    return url
            except requests.exceptions.ConnectionError:
                continue  # Domaine inexistant → continuer
            except requests.exceptions.Timeout:
                continue  # Timeout → continuer
            except Exception:
                continue
    except Exception:
        pass
    return None

def chercher_crt_sh(nom):
    """Cherche le domaine via les certificats SSL sur crt.sh."""
    try:
        # Prendre les 2-3 premiers mots du nom
        mots = nom.lower().split()[:2]
        query = " ".join(mots)
        url = f"https://crt.sh/?q={requests.utils.quote(query)}&output=json"
        resp = requete_humaine(url, timeout=8)
        if resp and resp.status_code == 200:
            data = resp.json()
            domaines = set()
            for cert in data[:20]:
                name = cert.get("name_value", "")
                for d in name.split("\n"):
                    d = d.strip().lstrip("*.")
                    if d and "." in d and len(d) < 50:
                        # Filtrer les domaines generiques
                        if not any(x in d for x in ["google", "cloudflare", "amazonaws", "microsoft"]):
                            domaines.add(d)
            # Tester chaque domaine trouvé
            for domaine in list(domaines)[:5]:
                url_test = f"https://{domaine}"
                try:
                    resp2 = requete_humaine(url_test, timeout=5)
                    if resp2 and resp2.status_code == 200:
                        # Vérifier que la page mentionne le nom de l'entreprise
                        if any(mot in resp2.text.lower() for mot in nom.lower().split()[:2]):
                            return url_test
                except Exception:
                    continue
    except Exception:
        pass
    return None

def chercher_viewdns_whois(nom_dirigeant):
    """Cherche les domaines enregistrés au nom du dirigeant via ViewDNS."""
    try:
        if not nom_dirigeant:
            return None
        url = f"https://viewdns.info/reversewhois/?q={requests.utils.quote(nom_dirigeant)}"
        soup = scraper_intelligent(url)
        if soup:
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if ".fr" in href or ".com" in href:
                    if href.startswith("http"):
                        return href
    except Exception:
        pass
    return None

def chercher_verif_com(siren):
    """Cherche sur verif.com via le SIREN."""
    try:
        if not siren:
            return None, "", ""
        url = f"https://www.verif.com/societe/{siren}"
        soup = scraper_intelligent(url)
        if soup:
            em, tel = extraire_depuis_soup(soup)
            # Chercher le site web
            site = ""
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if href.startswith("http") and "verif.com" not in href and "google" not in href:
                    site = href
                    break
            return site, em, tel
    except Exception:
        pass
    return None, "", ""

def chercher_societe_ninja(siren):
    """Cherche sur societe.ninja via le SIREN."""
    try:
        if not siren:
            return None, "", ""
        url = f"https://societe.ninja/{siren}"
        soup = scraper_intelligent(url)
        if soup:
            em, tel = extraire_depuis_soup(soup)
            site = ""
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if href.startswith("http") and "societe.ninja" not in href:
                    site = href
                    break
            return site, em, tel
    except Exception:
        pass
    return None, "", ""

def chercher_linkedin_entreprise(nom, ville):
    """Cherche la page LinkedIn de l'entreprise."""
    try:
        query = f"{nom} {ville} linkedin"
        url = f"https://www.google.com/search?q={requests.utils.quote(query)}"
        resp = requete_humaine(url, timeout=8)
        if resp:
            soup = BeautifulSoup(resp.text, "lxml")
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if "linkedin.com/company/" in href:
                    if not href.startswith("http"):
                        href = href.split("/url?q=")[1].split("&")[0] if "/url?q=" in href else href
                    return href
    except Exception:
        pass
    return None

# ─── RÉSEAUX SOCIAUX ─────────────────────────────────────────────────────────

# RESEAUX_SOCIAUX défini plus bas avec clés domaine/recherche

def extraire_contact_facebook(url):
    """Extrait email et téléphone depuis une page Facebook pro."""
    try:
        # Facebook bloque les scrapers, on essaie Camoufox en priorité
        soup = None
        if CAMOUFOX_DISPONIBLE:
            try:
                with Camoufox(headless=True) as fox:
                    page = fox.new_page()
                    page.set_extra_http_headers(get_headers_aleatoires())
                    page.goto(url, timeout=15000, wait_until="domcontentloaded")
                    time.sleep(random.uniform(2.0, 3.5))
                    # Aller sur la section "À propos"
                    about_url = url.rstrip("/") + "/about"
                    page.goto(about_url, timeout=10000, wait_until="domcontentloaded")
                    time.sleep(random.uniform(1.5, 2.5))
                    html = page.content()
                    page.close()
                    soup = BeautifulSoup(html, "lxml")
            except Exception:
                pass

        if not soup:
            resp = requete_humaine(url, timeout=8)
            if resp:
                soup = BeautifulSoup(resp.text, "lxml")

        if soup:
            texte = soup.get_text()
            emails = extraire_emails_complet(soup)
            tels = extraire_tels_complet(soup)
            em = emails[0][0] if emails else ""
            tel = tels[0][0] if tels else ""
            return em, tel
    except Exception:
        pass
    return "", ""

def extraire_contact_instagram(url):
    """Extrait email et téléphone depuis un profil Instagram pro."""
    try:
        # Essayer la version web basique
        resp = requete_humaine(url, timeout=8)
        if not resp:
            return "", ""
        soup = BeautifulSoup(resp.text, "lxml")
        texte = soup.get_text()
        # Instagram encode les données en JSON dans la page
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(script.string or "")
                if isinstance(data, list): data = data[0]
                em = data.get("email", "") or ""
                tel = data.get("telephone", "") or ""
                if em or tel:
                    return normaliser_email(em), normaliser_tel(tel) or ""
            except Exception:
                pass
        emails = extraire_emails_complet(soup)
        tels = extraire_tels_complet(soup)
        em = emails[0][0] if emails else ""
        tel = tels[0][0] if tels else ""
        return em, tel
    except Exception:
        pass
    return "", ""

def extraire_contact_linkedin(url):
    """Extrait email et téléphone depuis une page LinkedIn entreprise."""
    try:
        resp = requete_humaine(url, timeout=8)
        if not resp:
            return "", ""
        soup = BeautifulSoup(resp.text, "lxml")
        # LinkedIn met les infos dans des sections spécifiques
        for section in soup.find_all(["section", "div"],
                                      class_=re.compile(r"contact|about|details", re.I)):
            emails = extraire_emails_complet(section)
            tels = extraire_tels_complet(section)
            if emails or tels:
                em = emails[0][0] if emails else ""
                tel = tels[0][0] if tels else ""
                return em, tel
        emails = extraire_emails_complet(soup)
        tels = extraire_tels_complet(soup)
        em = emails[0][0] if emails else ""
        tel = tels[0][0] if tels else ""
        return em, tel
    except Exception:
        pass
    return "", ""

def extraire_contact_twitter(url):
    """Extrait email et téléphone depuis un profil X/Twitter."""
    try:
        resp = requete_humaine(url, timeout=8)
        if not resp:
            return "", ""
        soup = BeautifulSoup(resp.text, "lxml")
        # Chercher dans la bio / description
        for meta in soup.find_all("meta", {"name": re.compile(r"description|bio", re.I)}):
            contenu = meta.get("content", "")
            emails = [normaliser_email(e) for e, _ in deobfusquer_emails(contenu)]
            tels = extraire_tels_depuis_texte(contenu, "twitter_meta")
            if emails or tels:
                return emails[0] if emails else "", tels[0][0] if tels else ""
        emails = extraire_emails_complet(soup)
        tels = extraire_tels_complet(soup)
        em = emails[0][0] if emails else ""
        tel = tels[0][0] if tels else ""
        return em, tel
    except Exception:
        pass
    return "", ""

EXTRACTEURS_SOCIAUX = {
    "facebook": extraire_contact_facebook,
    "instagram": extraire_contact_instagram,
    "linkedin": extraire_contact_linkedin,
    "twitter": extraire_contact_twitter,
}

def chercher_similarweb(nom, ville):
    """Cherche sur SimilarWeb pour trouver le domaine."""
    try:
        query = f"{nom} {ville}"
        url = f"https://www.similarweb.com/fr/website-search/?query={requests.utils.quote(query)}"
        soup = scraper_intelligent(url)
        if soup:
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if "/website/" in href and "similarweb" not in href:
                    domaine = href.split("/website/")[-1].strip("/")
                    if domaine:
                        return f"https://{domaine}"
    except Exception:
        pass
    return None

def deviner_emails(site_web):
    """Génère et vérifie les emails probables via SMTP."""
    if not site_web:
        return ""
    try:
        domain = re.sub(r"https?://", "", site_web).split("/")[0].strip()
        if not domain:
            return ""
        candidats = [
            f"contact@{domain}", f"info@{domain}", f"bonjour@{domain}",
            f"accueil@{domain}", f"pro@{domain}", f"devis@{domain}",
        ]
        import smtplib as _smtp
        for email_candidat in candidats:
            try:
                mx_domain = domain
                server = _smtp.SMTP(timeout=5)
                server.connect("smtp.gmail.com", 25)
                server.helo("check.com")
                server.mail("check@check.com")
                code, _ = server.rcpt(email_candidat)
                server.quit()
                if code == 250:
                    return email_candidat
            except Exception:
                # Si SMTP bloqué, retourner le premier candidat probable
                return f"contact@{domain}"
    except Exception:
        pass
    return ""

# ─── RÉSEAUX SOCIAUX ─────────────────────────────────────────────────────────
RESEAUX_SOCIAUX = {
    "facebook": {
        "domaine": "facebook.com",
        "recherche": "https://www.google.com/search?q={query}+site:facebook.com",
    },
    "instagram": {
        "domaine": "instagram.com",
        "recherche": "https://www.google.com/search?q={query}+site:instagram.com",
    },
    "linkedin": {
        "domaine": "linkedin.com/company",
        "recherche": "https://www.google.com/search?q={query}+site:linkedin.com/company",
    },
    "twitter": {
        "domaine": "twitter.com",
        "recherche": "https://www.google.com/search?q={query}+site:twitter.com",
    },
}

def detecter_reseaux_sociaux(nom, ville, soup_homepage=None):
    """
    Détecte automatiquement les pages de réseaux sociaux d'une entreprise.
    Cherche d'abord dans la homepage, puis via Google.
    Limite à 3 réseaux maximum.
    """
    reseaux_trouves = {}

    # ── Source 1 : liens dans la homepage ─────────────────────────────────
    if soup_homepage:
        for a in soup_homepage.find_all("a", href=True):
            href = a["href"].lower()
            for reseau, config in RESEAUX_SOCIAUX.items():
                if reseau in reseaux_trouves:
                    continue
                if config["domaine"] in href:
                    url_full = a["href"]
                    if not url_full.startswith("http"):
                        url_full = "https://" + url_full.lstrip("/")
                    reseaux_trouves[reseau] = (url_full, "homepage")
                    break

    # ── Source 2 : Google si pas trouvé dans la homepage ──────────────────
    if len(reseaux_trouves) < 3:
        query = f"{nom} {ville}"
        for reseau, config in RESEAUX_SOCIAUX.items():
            if reseau in reseaux_trouves:
                continue
            if len(reseaux_trouves) >= 3:
                break
            try:
                url_recherche = config["recherche"].format(
                    query=requests.utils.quote(query)
                )
                delai_humain(0.5, 1.5)
                resp = requete_humaine(url_recherche, timeout=8)
                if not resp:
                    continue
                soup_google = BeautifulSoup(resp.text, "lxml")
                for a in soup_google.find_all("a", href=True):
                    href = a["href"]
                    if "/url?q=" in href:
                        real_url = href.split("/url?q=")[1].split("&")[0]
                        if config["domaine"] in real_url:
                            reseaux_trouves[reseau] = (real_url, "google")
                            break
            except Exception:
                continue

    return reseaux_trouves

def extraire_contacts_reseau_social(reseau, url):
    """
    Extrait emails et téléphones depuis un profil de réseau social.
    Analyse bio, description, informations publiques.
    """
    emails_trouves = []
    tels_trouves = []

    if not url:
        return emails_trouves, tels_trouves

    try:
        # Utiliser Camoufox/Playwright pour les réseaux sociaux
        # (souvent protégés par JS)
        if CAMOUFOX_DISPONIBLE or PLAYWRIGHT_DISPONIBLE:
            soup, emails_js = extraire_avec_playwright_js(url)
            if emails_js:
                emails_trouves.extend(emails_js)
            if soup:
                # Extraire depuis le HTML rendu
                ems = extraire_emails_complet(soup, url)
                emails_trouves.extend(ems)
                tels = extraire_tels_complet(soup)
                tels_trouves.extend(tels)
        else:
            # Fallback requests
            soup = scraper_intelligent(url)
            if soup:
                ems = extraire_emails_complet(soup, url)
                emails_trouves.extend(ems)
                tels = extraire_tels_complet(soup)
                tels_trouves.extend(tels)

        # Taguer la source avec le réseau social
        emails_trouves = [(em, f"{reseau}/{src}", conf)
                          for em, src, conf in emails_trouves]
        tels_trouves = [(tel, f"{reseau}/{src}", conf)
                        for tel, src, conf in tels_trouves]

    except Exception:
        pass

    return emails_trouves, tels_trouves

def rechercher_contacts_reseaux_sociaux(nom, ville, email_actuel="", tel_actuel="", soup_homepage=None):
    """
    Workflow complet :
    1. Détecter les réseaux sociaux (max 3)
    2. Pour chacun, extraire emails et téléphones
    3. Retourner le meilleur résultat
    """
    # Ne lancer que si email ou téléphone manquant
    if email_actuel and tel_actuel:
        return email_actuel, tel_actuel

    meilleur_email = email_actuel
    meilleur_tel = tel_actuel

    # Détecter les réseaux sociaux
    reseaux = detecter_reseaux_sociaux(nom, ville, soup_homepage)
    if not reseaux:
        return meilleur_email, meilleur_tel

    # Analyser chaque réseau (max 3, déjà limité par detecter_reseaux_sociaux)
    for reseau, (url_reseau, source_detection) in reseaux.items():
        if meilleur_email and meilleur_tel:
            break

        delai_humain(1.0, 2.5)
        emails_rs, tels_rs = extraire_contacts_reseau_social(reseau, url_reseau)

        if not meilleur_email and emails_rs:
            meilleur_email = emails_rs[0][0]

        if not meilleur_tel and tels_rs:
            meilleur_tel = tels_rs[0][0]

    return meilleur_email, meilleur_tel

def extraire_email_site(url):
    email, _ = extraire_contact_site(url)
    return email

def chercher_site_google(nom, ville):
    try:
        query = f"{nom} {ville} site officiel"
        url_s = f"https://www.google.com/search?q={requests.utils.quote(query)}"
        resp = requete_humaine(url_s, timeout=8)
        soup = BeautifulSoup(resp.text, "lxml")
        EXCLUS = ["google", "facebook", "pagesjaunes", "societe.com", "pappers", "infogreffe", "linkedin", "youtube", "wikipedia"]
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if "/url?q=" in href:
                real_url = href.split("/url?q=")[1].split("&")[0]
                if real_url.startswith("http") and not any(d in real_url for d in EXCLUS):
                    return real_url
    except Exception:
        pass
    return None

def chercher_pages_jaunes(nom, ville):
    """Cherche la fiche Pages Jaunes — essaie plusieurs formats d'URL."""
    headers_pj = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "fr-FR,fr;q=0.9",
        "Referer": "https://www.pagesjaunes.fr/",
    }
    query = f"{nom} {ville}"

    # Plusieurs formats d'URL Pages Jaunes
    urls_pj = [
        f"https://www.pagesjaunes.fr/annuaire/chercherlespros?quoiqui={requests.utils.quote(nom)}&ou={requests.utils.quote(ville)}",
        f"https://www.pagesjaunes.fr/annuaire/chercherlespros?quoiqui={requests.utils.quote(query)}",
        f"https://www.pagesjaunes.fr/recherche/sieste/univers/pros?quoiqui={requests.utils.quote(nom)}&ou={requests.utils.quote(ville)}",
    ]

    for url_s in urls_pj:
        try:
            resp = requests.get(url_s, timeout=10, headers=headers_pj, allow_redirects=True)
            if not resp or resp.status_code != 200:
                continue
            soup = BeautifulSoup(resp.text, "lxml")

            # Chercher les liens /pros/
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if "/pros/" in href:
                    if not href.startswith("http"):
                        href = "https://www.pagesjaunes.fr" + href
                    return href

            # Chercher via sélecteurs CSS variés
            for selector in ["bi-denomination", "denomination", "company-name",
                              "bi-names", "result-name", "NameEtab"]:
                lien = soup.find("a", class_=re.compile(selector, re.I))
                if lien and lien.get("href"):
                    href = lien["href"]
                    if not href.startswith("http"):
                        href = "https://www.pagesjaunes.fr" + href
                    return href

        except Exception:
            continue
    return None
def chercher_facebook(nom, ville):
    try:
        query = f"{nom} {ville} facebook"
        url_s = f"https://www.google.com/search?q={requests.utils.quote(query)}"
        resp = requete_humaine(url_s, timeout=8)
        soup = BeautifulSoup(resp.text, "lxml")
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if "/url?q=" in href:
                real_url = href.split("/url?q=")[1].split("&")[0]
                if "facebook.com" in real_url:
                    return real_url
    except Exception:
        pass
    return None

# ─── RECHERCHE ENTREPRISES ────────────────────────────────────────────────────
def geocode_adresse(adresse):
    """
    Géocode une adresse en coordonnées (lat, lon).
    Essaie d'abord l'API adresse.data.gouv.fr (fiable, sans quota),
    puis Nominatim en fallback.
    Gère les erreurs réseau sans planter l'application.
    """
    if not adresse:
        return None

    # ── 1. API adresse.data.gouv.fr (priorité — sans quota, fiable) ──────────
    try:
        url = f"https://api-adresse.data.gouv.fr/search/?q={requests.utils.quote(adresse)}&limit=1"
        resp = requests.get(url, timeout=8, headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code == 200:
            data = resp.json()
            features = data.get("features", [])
            if features:
                coords = features[0]["geometry"]["coordinates"]
                return (coords[1], coords[0])  # (lat, lon)
    except Exception:
        pass

    # ── 2. Nominatim en fallback ──────────────────────────────────────────────
    try:
        from geopy.geocoders import Nominatim
        from geopy.exc import GeocoderUnavailable, GeocoderTimedOut
        import uuid
        # User-Agent unique par session pour éviter les blocages
        ua = f"stage_tool_{uuid.uuid4().hex[:8]}"
        geolocator = Nominatim(user_agent=ua, timeout=10)
        location = geolocator.geocode(adresse)
        if location:
            return (location.latitude, location.longitude)
    except (GeocoderUnavailable, GeocoderTimedOut):
        pass
    except Exception:
        pass

    return None

def rechercher_entreprises(codes_ape, coords_ref, rayon_km, max_results, min_salaries, max_salaries, email_profil):
    entreprises = []
    historique = charger_historique(email_profil) if email_profil else {}
    for code_ape in codes_ape:
        code_ape = code_ape.strip().upper()
        page = 1
        while len(entreprises) < max_results:
            url = "https://recherche-entreprises.api.gouv.fr/search"
            params = {"activite_principale": code_ape, "per_page": 25, "page": page, "etat_administratif": "A"}
            try:
                resp = requests.get(url, params=params, timeout=10)
                data = resp.json()
            except Exception as e:
                st.warning(f"Erreur API pour {code_ape} : {e}")
                break
            resultats = data.get("results", [])
            if not resultats:
                break
            for r in resultats:
                try:
                    siege = r.get("siege", {})
                    lat = siege.get("latitude")
                    lon = siege.get("longitude")
                    if lat and lon:
                        dist = geodesic(coords_ref, (float(lat), float(lon))).km
                        if dist <= rayon_km:
                            tranche = r.get("tranche_effectif_salarie") or siege.get("tranche_effectif_salarie") or "00"
                            effectif = tranche_vers_effectif(tranche)
                            if effectif < min_salaries or effectif > max_salaries:
                                continue
                            nom = r.get("nom_complet", "")
                            cle = nom.strip().lower()
                            hist_entry = historique.get(cle, {})
                            entreprises.append({
                                "nom": nom,
                                "ville": siege.get("libelle_commune", ""),
                                "code_ape": r.get("activite_principale", ""),
                                "siren": r.get("siren", "") or r.get("numero_tva_intra", "")[:9] if r.get("numero_tva_intra") else r.get("siren", ""),
                                "adresse": siege.get("adresse", ""),
                                "site_web": hist_entry.get("site_web") or r.get("site_web", "") or "",
                                "pages_jaunes": hist_entry.get("pages_jaunes", ""),
                                "facebook": hist_entry.get("facebook", ""),
                                "linkedin": hist_entry.get("linkedin", ""),
                                "site_confidence": hist_entry.get("site_confidence", ""),
                                "site_validation_reason": hist_entry.get("site_validation_reason", ""),
                                "formulaires_contact": hist_entry.get("formulaires_contact", []),
                                "email": hist_entry.get("email", ""),
                                "telephone": hist_entry.get("telephone") or siege.get("telephone", "") or "",
                                "repondu": "Non", "j_ai_repondu": "Non",
                                "distance_km": round(dist, 1),
                                "effectif_approx": effectif,
                                "deja_contactee": cle in historique,
                                "date_contact": hist_entry.get("date_contact", ""),
                                "favori": hist_entry.get("favori", False),
                                "statut_kanban": hist_entry.get("statut_kanban", "📋 À contacter"),
                                "lat": float(lat), "lon": float(lon),
                            })
                except Exception:
                    continue
            if page >= data.get("total_pages", 1):
                break
            page += 1
            time.sleep(0.3)
    seen = set()
    unique = []
    for e in entreprises:
        if e["nom"] not in seen:
            seen.add(e["nom"])
            unique.append(e)
    return sorted(unique, key=lambda x: (not x["favori"], x["deja_contactee"], x["distance_km"]))

# ─── EXCEL ────────────────────────────────────────────────────────────────────
def generer_excel(entreprises, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidatures"
    headers = ["Nom", "Ville", "APE", "Répondu ?", "J'ai répondu ?", "Email", "Téléphone",
               "Adresse", "Distance (km)", "Effectif", "Déjà contactée ?", "Date contact",
               "Statut", "Favori ⭐", "Site web", "Pages Jaunes", "Facebook"]
    header_fill = PatternFill("solid", fgColor="2E86AB")
    grey_fill = PatternFill("solid", fgColor="DDDDDD")
    yellow_fill = PatternFill("solid", fgColor="FFF9C4")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    for row, e in enumerate(entreprises, 2):
        vals = [e["nom"], e["ville"], e["code_ape"], e.get("repondu","Non"), e.get("j_ai_repondu","Non"),
                e["email"], e["telephone"], e["adresse"], e["distance_km"], e.get("effectif_approx",""),
                "Oui" if e.get("deja_contactee") else "Non", e.get("date_contact",""),
                e.get("statut_kanban",""), "⭐" if e.get("favori") else "",
                e.get("site_web",""), e.get("pages_jaunes",""), e.get("facebook","")]
        fill = yellow_fill if e.get("favori") else (grey_fill if e.get("deja_contactee") else None)
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if fill:
                cell.fill = fill
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    wb.save(filepath)

# ─── VÉRIFICATION RELANCES ────────────────────────────────────────────────────
if st.session_state.profil_connecte and not st.session_state.relances_verifiees:
    p = st.session_state.profil_connecte
    if p.get("relance_auto", False):
        historique = charger_historique(p["email"])
        delai_jours = int(p.get("relance_delai", 7))
        nb_max_relances = int(p.get("relance_nb_max", 1))
        relances_en_attente = []
        for cle, e in historique.items():
            if e.get("repondu") or e.get("nb_relances", 0) >= nb_max_relances or not e.get("email"):
                continue
            date_ref_str = e.get("date_derniere_relance") or e.get("date_contact_iso", "")
            if not date_ref_str:
                continue
            try:
                date_ref = datetime.fromisoformat(date_ref_str)
                jours_ecoules = (datetime.now() - date_ref).days
                if jours_ecoules >= delai_jours:
                    relances_en_attente.append({**e, "jours_ecoules": jours_ecoules})
            except Exception:
                continue
        st.session_state.relances_en_attente = relances_en_attente
    st.session_state.relances_verifiees = True

# ─── INTERFACE ────────────────────────────────────────────────────────────────
st.title("🎓 Outil de Recherche de Stage")
st.markdown("---")

# ─── POPUP RELANCES ───────────────────────────────────────────────────────────
if st.session_state.profil_connecte and st.session_state.relances_en_attente:
    p = st.session_state.profil_connecte
    relances = st.session_state.relances_en_attente
    st.warning(f"🔔 **{len(relances)} relance(s) en attente !** Choisis lesquelles envoyer :")
    selections = {}
    for e in relances:
        selections[e["nom"]] = st.checkbox(
            f"**{e['nom']}** — {e.get('ville','')} — contactée il y a **{e['jours_ecoules']} jours**",
            value=True, key=f"relance_check_{e['nom'][:20]}"
        )
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📤 Envoyer les relances sélectionnées", type="primary"):
            a_envoyer = [e for e in relances if selections.get(e["nom"], False)]
            if not a_envoyer:
                st.warning("Aucune relance sélectionnée.")
            else:
                historique = charger_historique(p["email"])
                succes, echecs = 0, 0
                for e in a_envoyer:
                    sujet_p = p.get("sujet_relance", SUJET_RELANCE_DEFAULT).replace("{nom_entreprise}", e["nom"])
                    corps_p = p.get("corps_relance", CORPS_RELANCE_DEFAULT).replace("{nom_entreprise}", e["nom"])
                    corps_p = corps_p.replace("{ville}", e.get("ville","")).replace("{prenom_nom}", p.get("nom",""))
                    corps_p = corps_p.replace("{date_premier_contact}", e.get("date_contact",""))
                    result = envoyer_email(p["email"], p["mdp"], e["email"], sujet_p, corps_p)
                    if result is True:
                        cle = e["nom"].strip().lower()
                        if cle in historique:
                            historique[cle]["nb_relances"] = e.get("nb_relances", 0) + 1
                            historique[cle]["date_derniere_relance"] = datetime.now().isoformat()
                        succes += 1
                    else:
                        echecs += 1
                sauvegarder_historique(p["email"], historique)
                st.session_state.relances_en_attente = []
                st.success(f"✅ {succes} relance(s) envoyée(s) !")
                if echecs:
                    st.warning(f"⚠️ {echecs} échec(s).")
                st.rerun()
    with col2:
        if st.button("🚫 Ignorer pour cette session"):
            st.session_state.relances_en_attente = []
            st.rerun()
    st.markdown("---")

tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
    "👤 Profil", "🔍 Recherche", "📋 Kanban", "📧 Emails", "🗺️ Carte", "📈 Statistiques", "📊 Export Excel", "🔬 Logs"
])

# ══════════════════════════════════════════════════════════════════
# TAB 1 — PROFIL
# ══════════════════════════════════════════════════════════════════
with tab1:
    st.header("👤 Connexion à ton profil")
    st.info("Ton profil sauvegarde tous tes paramètres automatiquement.")

    col1, col2 = st.columns(2)
    with col1:
        profil_email_input = st.text_input("📧 Ton adresse Gmail", placeholder="toi@gmail.com", key="profil_email_input")
    with col2:
        profil_mdp_input = st.text_input("🔑 Mot de passe d'application", type="password", key="profil_mdp_input")

    if st.button("🔓 Se connecter / Créer le profil", type="primary"):
        if not profil_email_input or not profil_mdp_input:
            st.error("Remplis les deux champs !")
        else:
            profil_data = charger_profil(profil_email_input)
            historique = charger_historique(profil_email_input)
            st.session_state.profil_connecte = {"email": profil_email_input, "mdp": profil_mdp_input, **profil_data}
            # Cacher l'historique en session pour éviter les rechargements répétitifs
            st.session_state.historique_cache = historique
            st.session_state.relances_verifiees = False
            st.success(f"✅ Profil chargé pour **{profil_email_input}**")
            st.info(f"📬 **{len(historique)}** entreprise(s) dans ton historique.")

    if st.session_state.profil_connecte:
        p = st.session_state.profil_connecte
        st.markdown("---")
        st.subheader("⚙️ Mes paramètres")

        col1, col2 = st.columns(2)
        with col1:
            p_nom = st.text_input("👤 Prénom et nom", value=p.get("nom",""), placeholder="Jean Dupont", key="profil_nom")
            p_situation = st.text_input("🎓 Situation", value=p.get("situation",""), placeholder="En définition de projet", key="profil_situation")
            p_adresse = st.text_input("📍 Adresse de référence", value=p.get("adresse",""), placeholder="Richebourg, 62", key="profil_adresse")
            p_rayon = st.number_input("📏 Rayon (km)", min_value=1, max_value=100, value=int(p.get("rayon", 20)))
        with col2:
            p_codes_ape = st.text_area("🏭 Codes APE/NAF", value=p.get("codes_ape",""), placeholder="43.31Z\n81.30Z", height=120, key="profil_codes_ape")
            p_min_sal = st.number_input("👥 Salariés minimum", min_value=0, max_value=10000, value=int(p.get("min_sal", 1)))
            p_max_sal = st.number_input("👥 Salariés maximum", min_value=1, max_value=10000, value=int(p.get("max_sal", 50)))
            p_max_entreprises = st.number_input("🔢 Nombre max d'entreprises", min_value=10, max_value=500, value=int(p.get("max_entreprises", 100)))

        st.markdown("#### ✉️ Modèle email principal")
        p_sujet = st.text_input("Sujet", value=p.get("sujet","Demande de stage d'immersion – {nom_entreprise}"), key="profil_sujet")
        p_corps = st.text_area("Corps", value=p.get("corps", """Madame, Monsieur,

Je me permets de vous contacter afin de solliciter un stage d'immersion de deux semaines au sein de votre entreprise {nom_entreprise}, basée à {ville}.

Actuellement en définition de projet professionnel, je souhaite découvrir les métiers de votre secteur d'activité afin de mieux orienter mon parcours professionnel.

Sérieux(se), autonome et motivé(e), je suis prêt(e) à m'investir pleinement aux côtés de vos équipes. Ce stage ne représente aucune contrainte administrative ni financière pour votre entreprise.

Je reste disponible pour en discuter par téléphone ou par email, à votre convenance.

Dans l'attente de votre retour, je vous adresse mes sincères salutations.

{prenom_nom}"""), height=220, key="profil_corps")

        st.markdown("#### 📚 Modèles d'emails supplémentaires")
        st.caption("Crée des modèles par secteur et associe-leur des codes APE pour un envoi automatique.")
        modeles = charger_modeles(p["email"])
        if modeles:
            modele_a_voir = st.selectbox("Modèles existants", list(modeles.keys()), key="profil_modele_select")
            m = modeles[modele_a_voir]
            codes_assoc = [c.strip() for c in m.get("codes_ape","").strip().split("\n") if c.strip()]
            if codes_assoc:
                st.caption(f"🏭 Codes APE associés : {', '.join(codes_assoc)}")
            else:
                st.caption("🏭 Aucun code APE associé")
            col_a, col_b = st.columns(2)
            with col_b:
                if st.button("🗑️ Supprimer ce modèle", key="delete_modele"):
                    del modeles[modele_a_voir]
                    sauvegarder_modeles(p["email"], modeles)
                    st.success(f"Modèle supprimé !")
                    st.rerun()
            with col_a:
                if st.button("✏️ Modifier les codes APE", key="edit_ape_modele"):
                    st.session_state["edit_modele"] = modele_a_voir
            if st.session_state.get("edit_modele") == modele_a_voir:
                new_codes = st.text_area("Nouveaux codes APE", value=m.get("codes_ape",""), height=100, key="edit_ape_input")
                if st.button("💾 Mettre à jour", key="update_ape_modele"):
                    modeles[modele_a_voir]["codes_ape"] = new_codes
                    sauvegarder_modeles(p["email"], modeles)
                    st.session_state.profil_connecte["modeles_email"] = modeles
                    st.session_state.pop("edit_modele", None)
                    st.success("✅ Codes APE mis à jour !")
                    st.rerun()

        with st.expander("➕ Créer un nouveau modèle"):
            nouveau_nom = st.text_input("Nom du modèle", placeholder="Ex: BTP, Espaces verts...", key="nouveau_modele_nom")
            nouveau_sujet = st.text_input("Sujet", value="Demande de stage d'immersion – {nom_entreprise}", key="nouveau_modele_sujet")
            nouveau_corps = st.text_area("Corps", value="Madame, Monsieur,\n\n{prenom_nom}", height=180, key="nouveau_modele_corps")
            nouveau_codes_ape = st.text_area("🏭 Codes APE associés (un par ligne)", placeholder="43.31Z\n43.99C", height=100, key="nouveau_modele_ape")
            if st.button("💾 Sauvegarder ce modèle", key="save_modele"):
                if not nouveau_nom:
                    st.error("Donne un nom à ce modèle !")
                else:
                    modeles[nouveau_nom] = {"sujet": nouveau_sujet, "corps": nouveau_corps, "codes_ape": nouveau_codes_ape}
                    sauvegarder_modeles(p["email"], modeles)
                    st.session_state.profil_connecte["modeles_email"] = modeles
                    st.success(f"✅ Modèle '{nouveau_nom}' sauvegardé !")
                    st.rerun()

        st.markdown("#### 🔗 Associer un CV à des codes APE")
        st.caption("L'outil choisira automatiquement le bon CV selon le code APE de l'entreprise.")
        profil_frais = charger_profil(p["email"])
        cv_sauvegardes_profil = profil_frais.get("cv_noms", [])
        associations_ape = profil_frais.get("associations_ape_cv", {})
        st.session_state.profil_connecte["cv_noms"] = cv_sauvegardes_profil
        st.session_state.profil_connecte["associations_ape_cv"] = associations_ape

        if not cv_sauvegardes_profil:
            st.info("Aucun CV sauvegardé. Va dans l'onglet 📧 Emails pour uploader et sauvegarder un CV.")
        else:
            nouvelles_associations = {}
            for cv_nom in cv_sauvegardes_profil:
                codes_actuels = associations_ape.get(cv_nom, "")
                codes_input = st.text_area(f"📄 {cv_nom}", value=codes_actuels, placeholder="43.31Z\n43.99C", height=100, key=f"ape_cv_{cv_nom}")
                nouvelles_associations[cv_nom] = codes_input
            if st.button("💾 Sauvegarder les associations", key="save_ape_cv"):
                profil_data = charger_profil(p["email"])
                profil_data["associations_ape_cv"] = nouvelles_associations
                sauvegarder_profil(p["email"], profil_data)
                st.session_state.profil_connecte["associations_ape_cv"] = nouvelles_associations
                st.success("✅ Associations sauvegardées !")

        st.markdown("#### 🔔 Relances automatiques")
        p_relance_auto = st.toggle("Activer les relances automatiques", value=p.get("relance_auto", False))
        if p_relance_auto:
            col1, col2 = st.columns(2)
            with col1:
                p_relance_delai = st.number_input("⏳ Relancer après (jours)", min_value=1, max_value=30, value=int(p.get("relance_delai", 7)))
            with col2:
                p_relance_nb_max = st.number_input("🔁 Nombre max de relances", min_value=1, max_value=5, value=int(p.get("relance_nb_max", 1)))
            p_sujet_relance = st.text_input("Sujet relance", value=p.get("sujet_relance", SUJET_RELANCE_DEFAULT), key="profil_sujet_relance")
            p_corps_relance = st.text_area("Corps relance", value=p.get("corps_relance", CORPS_RELANCE_DEFAULT), height=200, key="profil_corps_relance")
        else:
            p_relance_delai = int(p.get("relance_delai", 7))
            p_relance_nb_max = int(p.get("relance_nb_max", 1))
            p_sujet_relance = p.get("sujet_relance", SUJET_RELANCE_DEFAULT)
            p_corps_relance = p.get("corps_relance", CORPS_RELANCE_DEFAULT)

        if st.button("💾 Sauvegarder mes paramètres", type="primary"):
            donnees = {
                "nom": p_nom, "situation": p_situation, "adresse": p_adresse,
                "rayon": p_rayon, "codes_ape": p_codes_ape,
                "min_sal": p_min_sal, "max_sal": p_max_sal, "max_entreprises": p_max_entreprises,
                "sujet": p_sujet, "corps": p_corps,
                "relance_auto": p_relance_auto, "relance_delai": p_relance_delai,
                "relance_nb_max": p_relance_nb_max, "sujet_relance": p_sujet_relance,
                "corps_relance": p_corps_relance,
                "cv_noms": cv_sauvegardes_profil,
                "associations_ape_cv": associations_ape,
                "modeles_email": charger_modeles(p["email"]),
            }
            sauvegarder_profil(p["email"], donnees)
            st.session_state.profil_connecte.update(donnees)
            st.success("✅ Paramètres sauvegardés !")

        st.markdown("---")
        st.subheader("📬 Historique")
        historique = get_historique_cached(p["email"])
        if not historique:
            st.write("Aucune entreprise contactée pour l'instant.")
        else:
            import pandas as pd
            recherche_hist = st.text_input("🔍 Rechercher dans l'historique", placeholder="Nom, ville...", key="search_historique")
            hist_filtre = {k: v for k, v in historique.items() if not recherche_hist or recherche_hist.lower() in v.get("nom","").lower() or recherche_hist.lower() in v.get("ville","").lower()}
            df_hist = pd.DataFrame([{
                "⭐": "⭐" if v.get("favori") else "", "Nom": v["nom"], "Ville": v["ville"],
                "Email": v["email"], "Date contact": v["date_contact"],
                "Statut": v.get("statut_kanban",""), "Relances": v.get("nb_relances",0),
                "Répondu": "✅" if v.get("repondu") else "❌"
            } for v in hist_filtre.values()])
            st.caption(f"{len(hist_filtre)} résultat(s) sur {len(historique)}")
            st.dataframe(df_hist, use_container_width=True, height=300)
            if st.button("🗑️ Réinitialiser l'historique", type="secondary"):
                sauvegarder_historique(p["email"], {})
                st.success("Historique réinitialisé !")
                st.rerun()

        st.markdown("---")
        st.subheader("🗂️ Sauvegarde & Restauration")
        col_exp, col_imp = st.columns(2)
        with col_exp:
            st.markdown("**📤 Exporter mes données**")
            if st.button("📤 Exporter", key="export_btn"):
                export_data = {
                    "profil": charger_profil(p["email"]),
                    "historique": charger_historique(p["email"]),
                    "email": p["email"],
                    "date_export": datetime.now().strftime("%d/%m/%Y %H:%M")
                }
                st.download_button(
                    label="⬇️ Télécharger la sauvegarde",
                    data=json.dumps(export_data, ensure_ascii=False, indent=2),
                    file_name=f"sauvegarde_stage_{datetime.now().strftime('%Y%m%d')}.json",
                    mime="application/json", key="download_export"
                )
        with col_imp:
            st.markdown("**📥 Importer une sauvegarde**")
            import_file = st.file_uploader("Fichier .json", type=["json"], key="import_file")
            if import_file:
                if st.button("📥 Restaurer", key="import_btn"):
                    try:
                        import_data = json.loads(import_file.read().decode("utf-8"))
                        if "profil" in import_data and "historique" in import_data:
                            sauvegarder_profil(p["email"], import_data["profil"])
                            sauvegarder_historique(p["email"], import_data["historique"])
                            st.session_state.profil_connecte.update(import_data["profil"])
                            st.success(f"✅ {len(import_data['historique'])} entreprises importées !")
                            st.rerun()
                        else:
                            st.error("Fichier invalide.")
                    except Exception as e:
                        st.error(f"Erreur : {e}")

# ══════════════════════════════════════════════════════════════════
# TAB 2 — RECHERCHE
# ══════════════════════════════════════════════════════════════════
with tab2:
    st.header("🔍 Rechercher des entreprises")
    if not st.session_state.profil_connecte:
        st.warning("⚠️ Connecte-toi d'abord dans l'onglet 👤 Profil.")
    else:
        p = st.session_state.profil_connecte
        col1, col2 = st.columns(2)
        with col1:
            adresse = st.text_input("📍 Adresse de référence", value=p.get("adresse",""), placeholder="Richebourg, 62", key="recherche_adresse")
            rayon = st.slider("📏 Rayon (km)", min_value=1, max_value=100, value=int(p.get("rayon", 20)))
        with col2:
            codes_ape_input = st.text_area("🏭 Codes APE/NAF (un par ligne)", value=p.get("codes_ape",""), height=120, key="recherche_codes_ape")
            max_results = st.number_input("Nombre max d'entreprises", min_value=10, max_value=500, value=int(p.get("max_entreprises", 100)))

        st.markdown("#### 👥 Filtre salariés")
        col3, col4 = st.columns(2)
        with col3:
            min_sal = st.number_input("Minimum", min_value=0, max_value=10000, value=int(p.get("min_sal", 1)), key="recherche_min_sal")
        with col4:
            max_sal = st.number_input("Maximum", min_value=1, max_value=10000, value=int(p.get("max_sal", 50)), key="recherche_max_sal")

        if st.button("🚀 Lancer la recherche", type="primary"):
            if not adresse or not codes_ape_input:
                st.error("Remplis l'adresse et au moins un code APE !")
            elif min_sal > max_sal:
                st.error("Le minimum ne peut pas être supérieur au maximum !")
            else:
                codes = [c for c in codes_ape_input.strip().split("\n") if c.strip()]
                with st.spinner("📍 Géolocalisation..."):
                    coords = geocode_adresse(adresse)
                if not coords:
                    st.error("Adresse introuvable.")
                else:
                    with st.spinner("🔎 Recherche en cours..."):
                        resultats = rechercher_entreprises(codes, coords, rayon, max_results, min_sal, max_sal, p["email"])
                    st.session_state.entreprises = resultats
                    nb_n = sum(1 for e in resultats if not e["deja_contactee"])
                    nb_d = sum(1 for e in resultats if e["deja_contactee"])
                    st.success(f"✅ {len(resultats)} entreprises — **{nb_n} nouvelles** / {nb_d} déjà contactées")

                    with st.spinner("🌐 Recherche automatique des sites web, emails et téléphones..."):

                        def chercher_site_complet(e):
                            """Cherche le site officiel via toutes les sources disponibles."""
                            if e.get("site_web"):
                                log_enrichissement(e["nom"], "site_web", f"Déjà connu : {e['site_web']}", "existant")
                                return

                            candidats = []

                            # 1. Tentative directe domaine
                            s = chercher_site_tentative_directe(e["nom"], e["ville"])
                            if s:
                                candidats.append((s, "domaine_direct"))
                                log_enrichissement(e["nom"], "candidat_site", s, "domaine_direct")

                            # 2. crt.sh
                            delai_humain(0.2, 0.5)
                            s = chercher_crt_sh(e["nom"])
                            if s:
                                candidats.append((s, "crt_sh"))
                                log_enrichissement(e["nom"], "candidat_site", s, "crt_sh")

                            # 3. Google Maps
                            delai_humain(0.2, 0.5)
                            _, _, s = chercher_google_maps(e["nom"], e["ville"])
                            if s:
                                candidats.append((s, "google_maps"))
                                log_enrichissement(e["nom"], "candidat_site", s, "google_maps")

                            # 4. Verif.com + Societe.ninja via SIREN
                            siren = e.get("siren", "")
                            if siren:
                                s, _, _ = chercher_verif_com(siren)
                                if s:
                                    candidats.append((s, "verif_com"))
                                    log_enrichissement(e["nom"], "candidat_site", s, "verif_com")
                                s, _, _ = chercher_societe_ninja(siren)
                                if s:
                                    candidats.append((s, "societe_ninja"))
                                    log_enrichissement(e["nom"], "candidat_site", s, "societe_ninja")

                            # 5. Mappy
                            s, _, _ = chercher_mappy(e["nom"], e["ville"])
                            if s:
                                candidats.append((s, "mappy"))
                                log_enrichissement(e["nom"], "candidat_site", s, "mappy")

                            # Scorer tous les candidats
                            if candidats:
                                meilleur, confidence, raison = trouver_meilleur_site(
                                    e["nom"], e["ville"], e.get("code_ape", ""), candidats
                                )
                                if meilleur:
                                    e["site_web"] = meilleur
                                    e["site_confidence"] = confidence
                                    e["site_validation_reason"] = raison
                                    log_enrichissement(e["nom"], "site_retenu", meilleur, confidence, rejete=False)
                                else:
                                    log_enrichissement(e["nom"], "site_rejete",
                                        f"{len(candidats)} candidats rejetés", raison, rejete=True)
                            else:
                                log_enrichissement(e["nom"], "site_introuvable", "Aucun candidat trouvé", rejete=True)

                        def chercher_contacts_complet(e, stats):
                            """Cherche email et téléphone via toutes les sources disponibles."""

                            def maj(em, tel, source=""):
                                if em and not e["email"]:
                                    if not est_email_hebergeur(em):
                                        e["email"] = em
                                        stats["email"] += 1
                                        log_enrichissement(e["nom"], "email_trouve", em, source)
                                    else:
                                        log_enrichissement(e["nom"], "email_rejete_hebergeur", em, source, rejete=True)
                                if tel and not e["telephone"]:
                                    if not est_tel_hebergeur(tel):
                                        e["telephone"] = tel
                                        stats["tel"] += 1
                                        log_enrichissement(e["nom"], "tel_trouve", tel, source)
                                    else:
                                        log_enrichissement(e["nom"], "tel_rejete_hebergeur", tel, source, rejete=True)

                            def ok():
                                return bool(e["email"] and e["telephone"])

                            # ── 1. API Annuaire Entreprises data.gouv.fr ──
                            if not ok():
                                em, tel = chercher_annuaire_gouv(e["nom"], e["ville"])
                                maj(em, tel, "annuaire_gouv")

                            # ── 2. Site web officiel (extraction + données structurées) ──
                            soup_site = None
                            if e.get("site_web") and not ok():
                                soup_site = scraper_intelligent(e["site_web"])
                                if soup_site:
                                    # Données structurées en priorité
                                    em_s, tel_s, _, _, _, _ = extraire_donnees_structurees(soup_site)
                                    maj(em_s, tel_s, "JSON-LD/schema.org")
                                    if not ok():
                                        em2, tel2 = extraire_depuis_soup(soup_site)
                                        maj(em2, tel2, "site_officiel")
                                    # Whois
                                    if not e["email"]:
                                        em_w = chercher_whois(e["site_web"])
                                        maj(em_w, "", "whois")
                                    # PDFs
                                    if not ok():
                                        em_p, tel_p, _, _ = analyser_pdfs_site(soup_site, e["site_web"])
                                        maj(em_p, tel_p, "pdf")

                            # ── 3. Google Maps ──
                            if not ok():
                                em, tel, _ = chercher_google_maps(e["nom"], e["ville"])
                                maj(em, tel, "google_maps")

                            # ── 4. Mappy ──
                            if not ok():
                                _, em, tel = chercher_mappy(e["nom"], e["ville"])
                                maj(em, tel, "mappy")

                            # ── 5. Pages Jaunes ──
                            if not ok() and e.get("pages_jaunes"):
                                em, tel = extraire_contact_site(e["pages_jaunes"])[:2]
                                maj(em, tel, "pages_jaunes")

                            # ── 6. Facebook ──
                            if not ok() and e.get("facebook"):
                                em, tel = extraire_contact_site(e["facebook"])[:2]
                                maj(em, tel, "facebook")

                            # ── 7. CMA ──
                            if not ok():
                                em, tel = chercher_cma(e["nom"], e["ville"])
                                maj(em, tel, "cma")

                            # ── 8. Qualibat ──
                            if not ok():
                                em, tel = chercher_qualibat(e["nom"], e["ville"])
                                maj(em, tel, "qualibat")

                            # ── 9. Faire.gouv.fr ──
                            if not ok():
                                em, tel = chercher_faire_gouv(e["nom"], e["ville"])
                                maj(em, tel, "faire_gouv")

                            # ── 10. Infobel ──
                            if not ok():
                                em, tel = chercher_infobel(e["nom"], e["ville"])
                                maj(em, tel, "infobel")

                            # ── 11. 118712 ──
                            if not ok():
                                em, tel = chercher_118712(e["nom"], e["ville"])
                                maj(em, tel, "118712")

                            # ── 12. Tel.fr ──
                            if not e["telephone"]:
                                em, tel = chercher_annuaire_tel(e["nom"], e["ville"])
                                maj(em, tel, "tel_fr")

                            # ── 13. Devis.fr ──
                            if not ok():
                                em, tel = chercher_devis_fr(e["nom"], e["ville"])
                                maj(em, tel, "devis_fr")

                            # ── 14. Pappers ──
                            if not ok():
                                em, tel = chercher_pappers(e["nom"], e["ville"])
                                maj(em, tel, "pappers")

                            # ── 15. Infogreffe ──
                            if not ok():
                                em, tel = chercher_infogreffe(e.get("siren", ""))
                                maj(em, tel, "infogreffe")

                            # ── 16. Societe.com ──
                            if not ok():
                                url_s = chercher_societe_com(e["nom"], e["ville"])
                                if url_s:
                                    em, tel = extraire_contact_site(url_s)[:2]
                                    maj(em, tel, "societe_com")

                            # ── 17. Cylex ──
                            if not ok():
                                em, tel = chercher_cylex(e["nom"], e["ville"])
                                maj(em, tel, "cylex")

                            # ── 18. Hoodspot ──
                            if not ok():
                                em, tel = chercher_hoodspot(e["nom"], e["ville"])
                                maj(em, tel, "hoodspot")

                            # ── 19. Europages ──
                            if not ok():
                                em, tel = chercher_europages(e["nom"], e["ville"])
                                maj(em, tel, "europages")

                            # ── 20. Kompass ──
                            if not ok():
                                url_k = chercher_kompass(e["nom"], e["ville"])
                                if url_k:
                                    em, tel = extraire_contact_site(url_k)[:2]
                                    maj(em, tel, "kompass")

                            # ── 21. Yelp ──
                            if not ok():
                                url_y = chercher_yelp(e["nom"], e["ville"])
                                if url_y:
                                    em, tel = extraire_contact_site(url_y)[:2]
                                    maj(em, tel, "yelp")

                            # ── 22. Verif.com ──
                            if not ok():
                                _, em, tel = chercher_verif_com(e.get("siren", ""))
                                maj(em, tel, "verif_com")

                            # ── 23. Réseaux sociaux ──
                            if not ok():
                                em_rs, tel_rs, _, _ = rechercher_contacts_reseaux_sociaux(
                                    e["nom"], e["ville"], soup_site, e.get("site_web", ""))
                                maj(em_rs, tel_rs, "reseau_social/")

                            # ── 24. OpenCorporates ──
                            if not ok():
                                em, tel = chercher_opencorporates(e["nom"], e["ville"])
                                maj(em, tel, "opencorporates")

                            # ── 25. Dun & Bradstreet ──
                            if not ok():
                                em, tel = chercher_dnb(e["nom"], e["ville"])
                                maj(em, tel, "dnb")

                            # ── 26. Google contact (dernier recours) ──
                            if not ok():
                                em, tel = chercher_google_contact(e["nom"], e["ville"])
                                maj(em, tel, "google")

                        # ── BOUCLE PRINCIPALE ────────────────────────────────
                        progress_auto = st.progress(0)
                        stats = {"site": 0, "pj": 0, "fb": 0, "email": 0, "tel": 0}
                        # Charger historique une seule fois pour tout le batch
                        hist_cache = charger_historique(p["email"])
                        hist_modifie = False

                        for i, e in enumerate(st.session_state.entreprises):
                            try:
                                # Chercher site web
                                chercher_site_complet(e)
                                if e.get("site_web"):
                                    stats["site"] += 1

                                # Chercher Pages Jaunes
                                if not e.get("pages_jaunes"):
                                    pj = chercher_pages_jaunes(e["nom"], e["ville"])
                                    if pj:
                                        e["pages_jaunes"] = pj
                                        stats["pj"] += 1

                                # Chercher Facebook
                                if not e.get("facebook"):
                                    fb = chercher_facebook(e["nom"], e["ville"])
                                    if fb:
                                        e["facebook"] = fb
                                        stats["fb"] += 1

                                # Chercher contacts
                                chercher_contacts_complet(e, stats)

                                # Formulaires de contact
                                if e.get("site_web") and not e.get("formulaires_contact"):
                                    try:
                                        soup_f = scraper_intelligent(e["site_web"], timeout=8)
                                        if soup_f:
                                            pages_f = decouvrir_pages_internes(e["site_web"], soup_f, max_pages=5)
                                            forms = chercher_formulaires_contact(e["site_web"], soup_f, pages_f)
                                            if forms:
                                                e["formulaires_contact"] = forms
                                    except Exception:
                                        pass

                                # Accumuler les mises à jour (batch, pas de save ici)
                                if e["deja_contactee"]:
                                    cle = e["nom"].strip().lower()
                                    if cle in hist_cache:
                                        hist_cache[cle].update({
                                            "site_web": e.get("site_web", ""),
                                            "pages_jaunes": e.get("pages_jaunes", ""),
                                            "facebook": e.get("facebook", ""),
                                            "email": e.get("email", ""),
                                            "telephone": e.get("telephone", ""),
                                            "formulaires_contact": e.get("formulaires_contact", []),
                                            "site_confidence": e.get("site_confidence", ""),
                                        })
                                        hist_modifie = True

                            except Exception as ex:
                                pass

                            progress_auto.progress((i + 1) / max(len(st.session_state.entreprises), 1))
                            delai_humain(1.0, 2.5)

                        # Sauvegarde batch une seule fois après la boucle
                        if hist_modifie:
                            sauvegarder_historique_cached(p["email"], hist_cache)

                        st.success(f"🌐 Sites : **{stats['site']}** | 📒 PJ : **{stats['pj']}** | 📘 FB : **{stats['fb']}** | 📧 Emails : **{stats['email']}** | 📞 Tél : **{stats['tel']}**")

        if st.session_state.entreprises:
            st.markdown("---")
            import pandas as pd
            col_f, col_d = st.columns(2)
            afficher_deja = col_d.checkbox("Afficher les déjà contactées", value=True)
            only_favoris = col_f.checkbox("⭐ Favoris uniquement", value=False)

            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("🕵️ Chercher emails + téléphones (manuel)"):
                    progress = st.progress(0)
                    for i, e in enumerate(st.session_state.entreprises):
                        for src in [e.get("site_web"), e.get("pages_jaunes"), e.get("facebook")]:
                            if src:
                                em, tel, _ = extraire_contact_site(src)
                                if em and not e["email"]: e["email"] = em
                                if tel and not e["telephone"]: e["telephone"] = tel
                            if e["email"] and e["telephone"]: break
                        progress.progress((i + 1) / len(st.session_state.entreprises))
                    st.success("Recherche de contacts terminée !")

            liste = st.session_state.entreprises
            if not afficher_deja:
                liste = [e for e in liste if not e["deja_contactee"]]
            if only_favoris:
                liste = [e for e in liste if e["favori"]]

            recherche_nom = st.text_input("🔍 Filtrer les résultats", placeholder="Nom, ville...", key="search_resultats")
            if recherche_nom:
                liste = [e for e in liste if recherche_nom.lower() in e["nom"].lower() or recherche_nom.lower() in e["ville"].lower()]
            st.caption(f"{len(liste)} entreprise(s) affichée(s)")

            df = pd.DataFrame([{
                "⭐": "⭐" if e["favori"] else "", "Nom": e["nom"], "Ville": e["ville"],
                "APE": e["code_ape"], "Dist. (km)": e["distance_km"],
                "Effectif": e.get("effectif_approx","?"), "Email": e["email"],
                "Téléphone": e["telephone"], "Statut": e.get("statut_kanban",""),
                "Contactée": "✅" if e["deja_contactee"] else "🆕",
            } for e in liste])
            st.dataframe(df, use_container_width=True, height=400)

            st.markdown("---")
            st.subheader("📇 Fiche entreprise")
            noms = [e["nom"] for e in st.session_state.entreprises]
            selected = st.selectbox("Choisir une entreprise", noms, key="select_modif")
            e_sel = next((e for e in st.session_state.entreprises if e["nom"] == selected), None)
            if e_sel:
                favori_icon = "⭐ " if e_sel.get("favori") else ""
                site_lien = f'<a href="{e_sel["site_web"]}" target="_blank">🌐 Voir</a>' if e_sel.get("site_web") else "🌐 Non trouvé"
                pj_lien = f'<a href="{e_sel["pages_jaunes"]}" target="_blank">📒 Voir</a>' if e_sel.get("pages_jaunes") else "📒 Non trouvé"
                fb_lien = f'<a href="{e_sel["facebook"]}" target="_blank">📘 Voir</a>' if e_sel.get("facebook") else "📘 Non trouvé"
                st.markdown(f"""
<div style="background:#1e1e3a;border-radius:12px;padding:20px;border-left:5px solid #4a90d9;margin-bottom:16px;">
<h3 style="margin:0 0 12px 0;color:#7eb8f7;">{favori_icon}{e_sel['nom']}</h3>
<table style="width:100%;border-collapse:collapse;color:#e2e8f0;">
<tr><td style="padding:4px 8px;color:#a0aec0;">📍 Adresse</td><td>{e_sel.get('adresse','Non renseignée')}</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">🏙️ Ville</td><td>{e_sel.get('ville','')}</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">🏭 Code APE</td><td>{e_sel.get('code_ape','')}</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">👥 Effectif</td><td>~{e_sel.get('effectif_approx','?')} salariés</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">📏 Distance</td><td>{e_sel.get('distance_km','?')} km</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">📧 Email</td><td>{e_sel.get('email','Non trouvé') or 'Non trouvé'}</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">📞 Téléphone</td><td>{e_sel.get('telephone','Non trouvé') or 'Non trouvé'}</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">🌐 Site web</td><td>{site_lien}</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">📒 Pages Jaunes</td><td>{pj_lien}</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">📘 Facebook</td><td>{fb_lien}</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">📋 Statut</td><td>{e_sel.get('statut_kanban','')}</td></tr>
<tr><td style="padding:4px 8px;color:#a0aec0;">📅 Contactée le</td><td>{e_sel.get('date_contact','Jamais') or 'Jamais'}</td></tr>
</table>
</div>""", unsafe_allow_html=True)
                with st.expander("✏️ Modifier cette fiche"):
                    col1, col2 = st.columns(2)
                    with col1:
                        fav = st.checkbox("⭐ Favori", value=e_sel.get("favori", False), key="modif_favori")
                        statut = st.selectbox("Statut Kanban", COLONNES_KANBAN,
                            index=COLONNES_KANBAN.index(e_sel.get("statut_kanban", COLONNES_KANBAN[0])) if e_sel.get("statut_kanban") in COLONNES_KANBAN else 0,
                            key="modif_statut")
                    with col2:
                        email_modif = st.text_input("📧 Email", value=e_sel.get("email",""), key="modif_email")
                        tel_modif = st.text_input("📞 Téléphone", value=e_sel.get("telephone",""), key="modif_tel")
                    col3, col4 = st.columns(2)
                    with col3:
                        site_modif = st.text_input("🌐 Site web", value=e_sel.get("site_web",""), key="modif_site")
                    with col4:
                        pj_modif = st.text_input("📒 Pages Jaunes", value=e_sel.get("pages_jaunes",""), key="modif_pj")
                    fb_modif = st.text_input("📘 Facebook", value=e_sel.get("facebook",""), key="modif_fb")
                    note = st.text_area("📝 Ajouter une note", value="", placeholder="Note horodatée...", key="modif_note", height=80)
                    if st.button("💾 Enregistrer les modifications"):
                        for e in st.session_state.entreprises:
                            if e["nom"] == selected:
                                e.update({"favori": fav, "statut_kanban": statut, "email": email_modif,
                                          "telephone": tel_modif, "site_web": site_modif,
                                          "pages_jaunes": pj_modif, "facebook": fb_modif})
                        hist = charger_historique(p["email"])
                        cle = selected.strip().lower()
                        if cle in hist:
                            hist[cle].update({"favori": fav, "statut_kanban": statut, "email": email_modif,
                                              "telephone": tel_modif, "site_web": site_modif,
                                              "pages_jaunes": pj_modif, "facebook": fb_modif})
                            if note:
                                hist[cle]["notes"] = (hist[cle].get("notes","") + f"\n[{datetime.now().strftime('%d/%m/%Y')}] {note}").strip()
                            sauvegarder_historique(p["email"], hist)
                        st.success("✅ Modifications enregistrées !")
                        st.rerun()

# ══════════════════════════════════════════════════════════════════
# TAB 3 — KANBAN
# ══════════════════════════════════════════════════════════════════
with tab3:
    st.header("📋 Tableau Kanban")
    if not st.session_state.profil_connecte:
        st.warning("⚠️ Connecte-toi d'abord dans l'onglet 👤 Profil.")
    else:
        p = st.session_state.profil_connecte
        historique = charger_historique(p["email"])
        if not historique:
            st.info("Aucune entreprise dans l'historique.")
        else:
            col_refresh, col_info = st.columns([1, 3])
            with col_refresh:
                auto_refresh = st.toggle("🔄 Actualisation auto (30s)", value=False, key="kanban_auto_refresh")
            if auto_refresh:
                st.markdown('<script>setTimeout(function(){window.location.reload();},30000);</script>', unsafe_allow_html=True)
            cols = st.columns(len(COLONNES_KANBAN))
            for i, col_name in enumerate(COLONNES_KANBAN):
                with cols[i]:
                    entreprises_col = [v for v in historique.values() if v.get("statut_kanban","📤 Contactée") == col_name]
                    auto_label = " 🔄" if i < 3 and auto_refresh else ""
                    st.markdown(f"**{col_name}{auto_label}** ({len(entreprises_col)})")
                    st.markdown("---")
                    for e in entreprises_col:
                        favori_icon = "⭐ " if e.get("favori") else ""
                        st.markdown(f"""<div style="background:#1e1e3a;border-radius:8px;padding:10px;margin-bottom:8px;border-left:4px solid #4a90d9;">
<b style="color:#e2e8f0;">{favori_icon}{e['nom']}</b><br>
<small style="color:#a0aec0;">📍 {e['ville']}</small><br>
<small style="color:#a0aec0;">📅 {e.get('date_contact','')}</small><br>
<small style="color:#a0aec0;">🔁 {e.get('nb_relances',0)} relance(s)</small>
</div>""", unsafe_allow_html=True)
                        nouveau_statut = st.selectbox("Déplacer vers", COLONNES_KANBAN,
                            index=COLONNES_KANBAN.index(col_name), key=f"kanban_{e['nom'][:20]}")
                        if nouveau_statut != col_name:
                            cle = e["nom"].strip().lower()
                            historique[cle]["statut_kanban"] = nouveau_statut
                            if nouveau_statut == "📬 Réponse reçue":
                                historique[cle]["repondu"] = True
                            sauvegarder_historique(p["email"], historique)
                            st.rerun()

# ══════════════════════════════════════════════════════════════════
# TAB 4 — EMAILS
# ══════════════════════════════════════════════════════════════════
with tab4:
    st.header("📧 Envoi automatique des emails")
    if not st.session_state.profil_connecte:
        st.warning("⚠️ Connecte-toi d'abord dans l'onglet 👤 Profil.")
    else:
        p = st.session_state.profil_connecte
        st.success(f"✅ Connecté : **{p['email']}**")

        with st.expander("📖 Comment créer un mot de passe d'application Gmail ?"):
            st.markdown("""
1. Va sur [myaccount.google.com](https://myaccount.google.com)
2. **Sécurité** → **Validation en deux étapes**
3. **Mots de passe des applications** → crée un mot de passe "Mail" → copie le code 16 caractères
            """)

        col1, col2 = st.columns(2)
        with col1:
            ton_nom = st.text_input("👤 Prénom et nom", value=p.get("nom",""), key="email_nom")
        with col2:
            ta_formation = st.text_input("🎓 Situation", value=p.get("situation",""), key="email_situation")

        st.markdown("---")
        st.subheader("📎 CV")
        profil_frais_cv = charger_profil(p["email"])
        cv_sauvegardes = profil_frais_cv.get("cv_noms", [])

        col_cv1, col_cv2 = st.columns(2)
        with col_cv1:
            cv_file = st.file_uploader("Ajouter / remplacer un CV (PDF)", type=["pdf"], key="cv_uploader")
            if cv_file:
                cv_label = st.text_input("Nom pour ce CV", placeholder="Ex: CV BTP, CV Espaces verts", key="cv_label_input")
                if st.button("💾 Sauvegarder ce CV", key="save_cv_btn"):
                    if not cv_label:
                        st.error("Donne un nom à ce CV !")
                    else:
                        chemin = f"cv_{get_profil_id(p['email'])}_{cv_label.replace(' ','_')}.pdf"
                        with open(chemin, "wb") as f_cv:
                            f_cv.write(cv_file.read())
                        profil_data = charger_profil(p["email"])
                        cv_liste = profil_data.get("cv_noms", [])
                        if cv_label not in cv_liste:
                            cv_liste.append(cv_label)
                        profil_data["cv_noms"] = cv_liste
                        sauvegarder_profil(p["email"], profil_data)
                        st.session_state.profil_connecte["cv_noms"] = cv_liste
                        st.success(f"✅ CV '{cv_label}' sauvegardé !")
                        st.rerun()
        with col_cv2:
            if cv_sauvegardes:
                cv_choisi_label = st.selectbox("📂 Choisir un CV sauvegardé", cv_sauvegardes, key="cv_select")
                chemin_cv_sauvegarde = f"cv_{get_profil_id(p['email'])}_{cv_choisi_label.replace(' ','_')}.pdf"
                if os.path.exists(chemin_cv_sauvegarde):
                    st.success(f"✅ CV sélectionné : **{cv_choisi_label}**")
                else:
                    st.warning("Ce CV n'existe plus sur le disque.")
                    chemin_cv_sauvegarde = None
                if st.button("🗑️ Supprimer ce CV", key="delete_cv_btn"):
                    if chemin_cv_sauvegarde and os.path.exists(chemin_cv_sauvegarde):
                        os.remove(chemin_cv_sauvegarde)
                    cv_sauvegardes.remove(cv_choisi_label)
                    profil_data = charger_profil(p["email"])
                    profil_data["cv_noms"] = cv_sauvegardes
                    sauvegarder_profil(p["email"], profil_data)
                    st.session_state.profil_connecte["cv_noms"] = cv_sauvegardes
                    st.rerun()
            else:
                st.info("Aucun CV sauvegardé.")
                chemin_cv_sauvegarde = None

        st.markdown("---")
        modeles_dispo = charger_modeles(p["email"])
        if modeles_dispo:
            col_mod1, col_mod2 = st.columns([2, 1])
            with col_mod1:
                modele_choisi = st.selectbox("📚 Utiliser un modèle", ["-- Modèle par défaut --"] + list(modeles_dispo.keys()), key="email_modele_select")
            with col_mod2:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("📥 Charger", key="load_modele"):
                    if modele_choisi != "-- Modèle par défaut --":
                        st.session_state["email_sujet_val"] = modeles_dispo[modele_choisi]["sujet"]
                        st.session_state["email_corps_val"] = modeles_dispo[modele_choisi]["corps"]
                        st.rerun()

        sujet_val = st.session_state.get("email_sujet_val", p.get("sujet","Demande de stage d'immersion – {nom_entreprise}"))
        corps_val = st.session_state.get("email_corps_val", p.get("corps",""))
        sujet = st.text_input("Sujet", value=sujet_val, key="email_sujet")
        corps = st.text_area("Corps", value=corps_val, height=280, key="email_corps")

        if p.get("relance_auto"):
            st.info(f"🔔 Relances auto activées : après **{p.get('relance_delai',7)} jours**, max **{p.get('relance_nb_max',1)} fois**.")

        st.markdown("---")
        if not st.session_state.entreprises:
            st.warning("⚠️ Lance d'abord une recherche dans l'onglet 🔍 Recherche")
        else:
            a_contacter = [e for e in st.session_state.entreprises if e["email"] and not e["deja_contactee"]]
            sans_email = [e for e in st.session_state.entreprises if not e["email"] and not e["deja_contactee"]]
            deja = [e for e in st.session_state.entreprises if e["deja_contactee"]]
            col1, col2, col3 = st.columns(3)
            col1.metric("🆕 Prêtes à envoyer", len(a_contacter))
            col2.metric("❌ Sans email", len(sans_email))
            col3.metric("✅ Déjà contactées", len(deja))

            if st.button("📤 Envoyer les emails", type="primary"):
                if not ton_nom:
                    st.error("Remplis ton prénom et nom !")
                elif not a_contacter:
                    st.warning("Aucune nouvelle entreprise avec un email.")
                else:
                    chemin_cv_temp = None
                    if cv_file:
                        chemin_cv_temp = f"cv_temp_{cv_file.name}"
                        with open(chemin_cv_temp, "wb") as f_tmp:
                            f_tmp.write(cv_file.read())

                    associations_ape_cv = charger_profil(p["email"]).get("associations_ape_cv", {})
                    modeles_send = charger_modeles(p["email"])

                    def get_cv_pour_entreprise(code_ape):
                        for cv_nom, codes_str in associations_ape_cv.items():
                            codes = [c.strip().upper() for c in codes_str.strip().split("\n") if c.strip()]
                            if code_ape.upper() in codes:
                                chemin = f"cv_{get_profil_id(p['email'])}_{cv_nom.replace(' ','_')}.pdf"
                                if os.path.exists(chemin):
                                    return chemin, cv_nom
                        if chemin_cv_sauvegarde and os.path.exists(chemin_cv_sauvegarde):
                            return chemin_cv_sauvegarde, cv_choisi_label
                        if chemin_cv_temp:
                            return chemin_cv_temp, "CV temporaire"
                        return None, None

                    def get_modele_pour_entreprise(code_ape):
                        for modele_nom, modele_data in modeles_send.items():
                            codes = [c.strip().upper() for c in modele_data.get("codes_ape","").strip().split("\n") if c.strip()]
                            if code_ape.upper() in codes:
                                return modele_data["sujet"], modele_data["corps"], modele_nom
                        return sujet, corps, "Modèle par défaut"

                    progress = st.progress(0)
                    succes_list, echecs = [], 0
                    cv_utilises, modeles_utilises = {}, {}

                    for i, e in enumerate(a_contacter):
                        sujet_auto, corps_auto, nom_modele = get_modele_pour_entreprise(e.get("code_ape",""))
                        corps_p = corps_auto.replace("{nom_entreprise}", e["nom"]).replace("{ville}", e["ville"])
                        corps_p = corps_p.replace("{prenom_nom}", ton_nom).replace("{formation}", ta_formation)
                        sujet_p = sujet_auto.replace("{nom_entreprise}", e["nom"])
                        chemin_cv, nom_cv = get_cv_pour_entreprise(e.get("code_ape",""))
                        cv_utilises[nom_cv or "Aucun"] = cv_utilises.get(nom_cv or "Aucun", 0) + 1
                        modeles_utilises[nom_modele] = modeles_utilises.get(nom_modele, 0) + 1
                        result = envoyer_email(p["email"], p["mdp"], e["email"], sujet_p, corps_p, chemin_cv)
                        if result is True:
                            succes_list.append(e)
                        else:
                            echecs += 1
                        progress.progress((i + 1) / len(a_contacter))
                        time.sleep(1)

                    if succes_list:
                        marquer_contactees(p["email"], succes_list)
                        for e in st.session_state.entreprises:
                            if e in succes_list:
                                e["deja_contactee"] = True
                                e["date_contact"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                                e["statut_kanban"] = "📤 Contactée"

                    if chemin_cv_temp and os.path.exists(chemin_cv_temp):
                        os.remove(chemin_cv_temp)

                    st.success(f"✅ {len(succes_list)} emails envoyés !")
                    if echecs:
                        st.warning(f"⚠️ {echecs} échecs.")
                    if cv_utilises:
                        st.info("📄 CVs : " + " | ".join([f"**{k}** : {v}" for k, v in cv_utilises.items()]))
                    if modeles_utilises:
                        st.info("✉️ Modèles : " + " | ".join([f"**{k}** : {v}" for k, v in modeles_utilises.items()]))

# ══════════════════════════════════════════════════════════════════
# TAB 5 — CARTE
# ══════════════════════════════════════════════════════════════════
with tab5:
    st.header("🗺️ Carte des entreprises")
    if not st.session_state.entreprises:
        st.warning("⚠️ Lance d'abord une recherche.")
    else:
        try:
            import folium
            from streamlit_folium import st_folium
            liste = st.session_state.entreprises
            coords_valides = [e for e in liste if e.get("lat") and e.get("lon")]
            if not coords_valides:
                st.warning("Aucune coordonnée disponible.")
            else:
                center_lat = sum(e["lat"] for e in coords_valides) / len(coords_valides)
                center_lon = sum(e["lon"] for e in coords_valides) / len(coords_valides)
                m = folium.Map(location=[center_lat, center_lon], zoom_start=11)
                for e in coords_valides:
                    color = "gold" if e.get("favori") else ("green" if e.get("deja_contactee") else "blue")
                    icon = "star" if e.get("favori") else ("check" if e.get("deja_contactee") else "info-sign")
                    popup_html = f"<b>{e['nom']}</b><br>📍 {e['ville']}<br>📏 {e['distance_km']} km<br>📧 {e['email'] or 'Pas d\'email'}<br>📞 {e['telephone'] or 'Pas de téléphone'}"
                    folium.Marker(location=[e["lat"], e["lon"]], popup=folium.Popup(popup_html, max_width=250),
                        tooltip=e["nom"], icon=folium.Icon(color=color, icon=icon, prefix="glyphicon")).add_to(m)
                st.caption("🔵 Nouvelle  🟢 Déjà contactée  🟡 Favori")
                st_folium(m, width=None, height=550, use_container_width=True)
        except ImportError:
            st.info("Pour activer la carte, tape dans le terminal :")
            st.code("pip install folium streamlit-folium")

# ══════════════════════════════════════════════════════════════════
# TAB 6 — STATISTIQUES
# ══════════════════════════════════════════════════════════════════
with tab6:
    st.header("📈 Tableau de bord")
    if not st.session_state.profil_connecte:
        st.warning("⚠️ Connecte-toi d'abord dans l'onglet 👤 Profil.")
    else:
        p = st.session_state.profil_connecte
        historique = charger_historique(p["email"])
        if not historique:
            st.info("Aucune donnée disponible.")
        else:
            import pandas as pd
            total = len(historique)
            repondus = sum(1 for e in historique.values() if e.get("repondu"))
            favoris = sum(1 for e in historique.values() if e.get("favori"))
            entretiens = sum(1 for e in historique.values() if e.get("statut_kanban") == "🤝 Entretien")
            acceptes = sum(1 for e in historique.values() if e.get("statut_kanban") == "✅ Accepté")
            relances_total = sum(e.get("nb_relances", 0) for e in historique.values())
            taux_reponse = round(repondus / total * 100, 1) if total > 0 else 0

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("📤 Emails envoyés", total)
            col2.metric("📬 Réponses", repondus, f"{taux_reponse}%")
            col3.metric("🤝 Entretiens", entretiens)
            col4.metric("✅ Acceptés", acceptes)
            col5, col6, col7 = st.columns(3)
            col5.metric("⭐ Favoris", favoris)
            col6.metric("🔁 Relances", relances_total)
            col7.metric("📊 Taux réponse", f"{taux_reponse}%")

            st.markdown("---")
            st.subheader("📊 Répartition par statut")
            statuts = {}
            for e in historique.values():
                s = e.get("statut_kanban","📤 Contactée")
                statuts[s] = statuts.get(s, 0) + 1
            st.bar_chart(pd.DataFrame(list(statuts.items()), columns=["Statut","Nombre"]).set_index("Statut"))

            st.subheader("🏙️ Top villes")
            villes = {}
            for e in historique.values():
                v = e.get("ville","Inconnue")
                villes[v] = villes.get(v, 0) + 1
            df_villes = pd.DataFrame(list(villes.items()), columns=["Ville","Nombre"]).sort_values("Nombre", ascending=False).head(10)
            st.bar_chart(df_villes.set_index("Ville"))

# ══════════════════════════════════════════════════════════════════
# TAB 7 — EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════
with tab7:
    st.header("📊 Export Excel")
    if not st.session_state.entreprises:
        st.warning("⚠️ Lance d'abord une recherche dans l'onglet 🔍 Recherche")
    else:
        nb_n = sum(1 for e in st.session_state.entreprises if not e["deja_contactee"])
        nb_d = sum(1 for e in st.session_state.entreprises if e["deja_contactee"])
        nb_f = sum(1 for e in st.session_state.entreprises if e["favori"])
        col1, col2, col3 = st.columns(3)
        col1.metric("🆕 Nouvelles", nb_n)
        col2.metric("✅ Déjà contactées", nb_d)
        col3.metric("⭐ Favoris", nb_f)
        st.caption("Les favoris sont en jaune, les déjà contactées en gris dans le fichier.")
        filepath = "candidatures_stage.xlsx"
        if st.button("📥 Générer le fichier Excel", type="primary"):
            generer_excel(st.session_state.entreprises, filepath)
            with open(filepath, "rb") as f:
                st.download_button(
                    label="⬇️ Télécharger candidatures_stage.xlsx",
                    data=f,
                    file_name="candidatures_stage.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            st.success("✅ Fichier Excel généré !")

# ══════════════════════════════════════════════════════════════════
# TAB 8 — LOGS
# ══════════════════════════════════════════════════════════════════
with tab8:
    st.header("🔬 Logs d'enrichissement")
    st.caption("Suivi détaillé de chaque recherche : domaines candidats, scores, emails/téléphones trouvés, rejets.")

    logs = get_logs()
    if not logs:
        st.info("Aucun log disponible. Lance une recherche pour voir les logs.")
    else:
        col1, col2 = st.columns(2)
        with col1:
            filtre_entreprise = st.text_input("🔍 Filtrer par entreprise", key="log_filter")
        with col2:
            filtre_action = st.selectbox("Filtrer par action", 
                ["Tous", "site_retenu", "site_rejete", "email_trouve", "tel_trouve",
                 "email_rejete_hebergeur", "tel_rejete_hebergeur", "site_introuvable"],
                key="log_action_filter")

        logs_filtres = logs
        if filtre_entreprise:
            logs_filtres = [l for l in logs_filtres if filtre_entreprise.lower() in l["entreprise"].lower()]
        if filtre_action != "Tous":
            logs_filtres = [l for l in logs_filtres if l["action"] == filtre_action]

        st.caption(f"{len(logs_filtres)} entrée(s) sur {len(logs)} total")

        import pandas as pd
        df_logs = pd.DataFrame([{
            "⏱️": l["ts"],
            "Entreprise": l["entreprise"],
            "Action": l["action"],
            "Détail": l["detail"],
            "Source": l["source"],
            "Score": l["score"] or "",
            "❌ Rejeté": "❌" if l["rejete"] else "",
        } for l in reversed(logs_filtres)])
        st.dataframe(df_logs, use_container_width=True, height=500)

        if st.button("🗑️ Effacer les logs"):
            clear_logs()
            st.rerun()
